import os
import sys
import json
import uuid
import pathlib
import anthropic
from dotenv import load_dotenv

load_dotenv()

import asyncio
from fastapi import FastAPI, HTTPException, Header, UploadFile, Body, File, Form, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from typing import Literal
import sqlite3
import time
from datetime import datetime, timezone
from contextlib import contextmanager

DB_PATH = os.path.join(os.path.dirname(__file__), "admin.db")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "admin123")  # 建议在 .env 中设置强密码

import hashlib
import secrets

def hash_password(pwd: str) -> str:
    return hashlib.sha256(pwd.encode()).hexdigest()

def init_db():
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                username   TEXT NOT NULL UNIQUE,
                password   TEXT NOT NULL,
                role       TEXT DEFAULT 'user',
                created_at TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS user_tokens (
                token      TEXT PRIMARY KEY,
                user_id    INTEGER REFERENCES users(id),
                created_at TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS api_requests (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id      INTEGER REFERENCES users(id),
                config_id    INTEGER REFERENCES api_configs(id),
                project_name TEXT NOT NULL,
                purpose      TEXT NOT NULL,
                lead         TEXT NOT NULL,
                budget       TEXT NOT NULL,
                sub_accounts TEXT DEFAULT '',
                status       TEXT DEFAULT 'pending',
                review_note  TEXT DEFAULT '',
                created_at   TEXT NOT NULL,
                updated_at   TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS api_configs (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                name       TEXT NOT NULL,
                base_url   TEXT NOT NULL,
                api_key    TEXT NOT NULL,
                provider   TEXT NOT NULL,
                models     TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                is_active  INTEGER DEFAULT 1
            )
        """)
        # 兼容旧库：新增字段
        for col, definition in [
            ("models",          "TEXT DEFAULT ''"),
            ("price_input",     "REAL DEFAULT 0"),   # 每千 input token 价格（元）
            ("price_output",    "REAL DEFAULT 0"),   # 每千 output token 价格（元）
            ("manager",         "TEXT DEFAULT ''"),  # 该 API 的负责管理员
        ]:
            try:
                conn.execute(f"ALTER TABLE api_configs ADD COLUMN {col} {definition}")
                conn.commit()
            except Exception:
                pass
        conn.execute("""
            CREATE TABLE IF NOT EXISTS usage_stats (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                config_id     INTEGER REFERENCES api_configs(id),
                user_id       INTEGER REFERENCES users(id),
                called_at     TEXT NOT NULL,
                model         TEXT,
                input_tokens  INTEGER DEFAULT 0,
                output_tokens INTEGER DEFAULT 0,
                cost          REAL DEFAULT 0,
                success       INTEGER DEFAULT 1,
                duration_ms   INTEGER DEFAULT 0,
                error_msg     TEXT
            )
        """)
        # 兼容旧 usage_stats
        for col, definition in [
            ("user_id", "INTEGER DEFAULT NULL"),
            ("cost",    "REAL DEFAULT 0"),
        ]:
            try:
                conn.execute(f"ALTER TABLE usage_stats ADD COLUMN {col} {definition}")
                conn.commit()
            except Exception:
                pass
        conn.execute("""
            CREATE TABLE IF NOT EXISTS chat_sessions (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id    INTEGER REFERENCES users(id),
                config_id  INTEGER REFERENCES api_configs(id),
                model      TEXT,
                title      TEXT DEFAULT '',
                created_at TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS chat_messages (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                session_id INTEGER REFERENCES chat_sessions(id),
                role       TEXT NOT NULL,
                content    TEXT NOT NULL,
                created_at TEXT NOT NULL
            )
        """)
        # Spec 1 三层结构(accounts → sub_accounts → api_keys)
        # 生产中由 migrations/v1_account_schema.py 创建;tests 用 init_db 兜底
        conn.execute("""
            CREATE TABLE IF NOT EXISTS accounts (
                id                   INTEGER PRIMARY KEY AUTOINCREMENT,
                provider             TEXT NOT NULL,
                base_url             TEXT NOT NULL,
                provider_backend_url TEXT DEFAULT '',
                quota_total_path     TEXT DEFAULT '',
                balance_path         TEXT DEFAULT '',
                cost_path            TEXT DEFAULT '',
                manager_user_id      INTEGER REFERENCES users(id),
                team                 TEXT DEFAULT '',
                created_by           INTEGER NOT NULL REFERENCES users(id),
                models               TEXT DEFAULT '',
                is_active            INTEGER DEFAULT 1,
                created_at           TEXT NOT NULL,
                updated_at           TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS sub_accounts (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                account_id  INTEGER NOT NULL REFERENCES accounts(id) ON DELETE CASCADE,
                name        TEXT NOT NULL,
                description TEXT DEFAULT '',
                created_at  TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS api_keys (
                id              INTEGER PRIMARY KEY,
                sub_account_id  INTEGER NOT NULL REFERENCES sub_accounts(id) ON DELETE RESTRICT,
                name            TEXT NOT NULL,
                api_key         TEXT NOT NULL,
                is_active       INTEGER DEFAULT 1,
                exhausted       INTEGER DEFAULT 0,
                created_at      TEXT NOT NULL
            )
        """)
        # Spec 2: api_keys quota cache + 3-strike exhausted 计数
        for col, defn in [
            ("last_total",   "REAL"),
            ("last_balance", "REAL"),
            ("last_used",    "REAL"),
            ("last_quota_at","TEXT"),
            ("zero_count",   "INTEGER DEFAULT 0"),
            ("last_zero_at", "TEXT"),
            ("manager_user_id", "INTEGER REFERENCES users(id)"),
            ("created_by",   "INTEGER REFERENCES users(id)"),
        ]:
            try:
                conn.execute(f"ALTER TABLE api_keys ADD COLUMN {col} {defn}")
                conn.commit()
            except Exception:
                pass
        # Spec 2: accounts 上每个接口的 JSON 提取路径
        for col, defn in [
            ("quota_total_json_path", "TEXT DEFAULT ''"),
            ("balance_json_path",     "TEXT DEFAULT ''"),
            ("cost_json_path",        "TEXT DEFAULT ''"),
        ]:
            try:
                conn.execute(f"ALTER TABLE accounts ADD COLUMN {col} {defn}")
                conn.commit()
            except Exception:
                pass
        # chat_sessions 兼容添加 pinned 字段
        try:
            conn.execute("ALTER TABLE chat_sessions ADD COLUMN pinned INTEGER DEFAULT 0")
            conn.commit()
        except Exception:
            pass
        # chat_messages 兼容添加 model 字段
        try:
            conn.execute("ALTER TABLE chat_messages ADD COLUMN model TEXT DEFAULT ''")
            conn.commit()
        except Exception:
            pass
        # batch_jobs / batch_job_rows（SQLite 批量历史，不依赖 ClickHouse）
        conn.execute("""
            CREATE TABLE IF NOT EXISTS batch_jobs (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id    INTEGER REFERENCES users(id),
                task_name  TEXT NOT NULL,
                model      TEXT DEFAULT '',
                config_id  INTEGER,
                label      TEXT DEFAULT '',
                row_count  INTEGER DEFAULT 0,
                done_count INTEGER DEFAULT 0,
                fail_count INTEGER DEFAULT 0,
                status     TEXT DEFAULT 'running',
                created_at TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS batch_job_rows (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                job_id      INTEGER REFERENCES batch_jobs(id),
                row_index   INTEGER NOT NULL,
                input_json  TEXT NOT NULL,
                output_text TEXT DEFAULT '',
                success     INTEGER DEFAULT 1,
                error_msg   TEXT DEFAULT ''
            )
        """)
        # batch_job_rows 兼容添加字段
        for col, definition in [
            ("output_type",  "TEXT DEFAULT 'text'"),
            ("output_path",  "TEXT DEFAULT ''"),
            ("label",        "TEXT DEFAULT ''"),
            ("started_at",   "TEXT DEFAULT ''"),
            ("finished_at",  "TEXT DEFAULT ''"),
        ]:
            try:
                conn.execute(f"ALTER TABLE batch_job_rows ADD COLUMN {col} {definition}")
            except Exception:
                pass
        # batch_jobs 兼容添加字段
        for col, definition in [
            ("settings_json", "TEXT DEFAULT ''"),
            ("batch_id",      "TEXT DEFAULT ''"),
            ("started_at",    "TEXT DEFAULT ''"),
            ("finished_at",   "TEXT DEFAULT ''"),
            ("task_id",       "INTEGER DEFAULT NULL"),
            ("source_type",   "TEXT DEFAULT 'click'"),
            ("script_code",   "TEXT DEFAULT ''"),
            ("config_json",   "TEXT DEFAULT '{}'"),
        ]:
            try:
                conn.execute(f"ALTER TABLE batch_jobs ADD COLUMN {col} {definition}")
            except Exception:
                pass

        # tasks 表（任务 = 配置+脚本的容器；一个任务可有多次 run/batch_job）
        conn.execute("""
            CREATE TABLE IF NOT EXISTS tasks (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id         INTEGER REFERENCES users(id),
                task_name       TEXT NOT NULL,
                config_json     TEXT DEFAULT '{}',
                script_code     TEXT DEFAULT '',
                script_is_dirty INTEGER DEFAULT 0,
                created_at      TEXT NOT NULL,
                updated_at      TEXT NOT NULL
            )
        """)
        # api_requests 兼容添加字段
        for col, definition in [
            ("sub_account_id", "INTEGER DEFAULT NULL"),
            ("cc_person",      "TEXT DEFAULT ''"),
            ("account_id",     "INTEGER DEFAULT NULL"),
            ("api_key_id",     "INTEGER DEFAULT NULL"),
            ("dept",           "TEXT DEFAULT ''"),
            ("reviewer_id",    "INTEGER DEFAULT NULL"),
        ]:
            try:
                conn.execute(f"ALTER TABLE api_requests ADD COLUMN {col} {definition}")
            except Exception:
                pass
        conn.commit()
        # 初始化默认管理员账号（仅首次，已存在则跳过）
        now = datetime.now(timezone.utc).isoformat()
        try:
            conn.execute(
                "INSERT INTO users (username, password, role, created_at) VALUES (?,?,?,?)",
                ("admin", hash_password("admin123"), "admin", now)
            )
            conn.commit()
            print("INFO: 默认管理员账号已创建 (admin / admin123)")
        except Exception:
            pass  # 已存在则跳过

init_db()

@contextmanager
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
    finally:
        conn.close()

def require_admin(x_admin_password: str = Header(default="")):
    import hmac
    if not hmac.compare_digest(x_admin_password, ADMIN_PASSWORD):
        raise HTTPException(status_code=403, detail="Forbidden")

_env_api_key = os.environ.get("ANTHROPIC_API_KEY")
if not _env_api_key:
    print("WARNING: ANTHROPIC_API_KEY not set. Will use key from admin database.")

def get_active_anthropic_key() -> str | None:
    """从 api_configs 表读取 provider=anthropic 且 is_active=1 的最新密钥（兜底用）。"""
    try:
        with get_db() as conn:
            row = conn.execute(
                "SELECT api_key FROM api_configs WHERE provider='anthropic' AND is_active=1 ORDER BY id DESC LIMIT 1"
            ).fetchone()
        return row["api_key"] if row else None
    except Exception:
        return None

def get_config_credentials(config_id: int | None) -> tuple[str | None, str | None, str | None]:
    """返回 (api_key, base_url, provider)。优先用指定 config_id，否则兜底找 anthropic provider。"""
    if config_id:
        try:
            with get_db() as conn:
                row = conn.execute(
                    "SELECT api_key, base_url, provider FROM api_configs WHERE id=? AND is_active=1",
                    (config_id,)
                ).fetchone()
            if row:
                return row["api_key"], row["base_url"], row["provider"]
        except Exception:
            pass
    return get_active_anthropic_key() or _env_api_key, None, "anthropic"

def _friendly_error(e: Exception) -> str:
    """把 API 原始异常转成简洁的中文提示。"""
    raw = str(e)
    import re as _re
    # 提取 message 字段
    m = _re.search(r"'message':\s*'([^']+)'", raw)
    if not m:
        m = _re.search(r'"message":\s*"([^"]+)"', raw)
    if m:
        msg = m.group(1)
        # 常见错误的友好提示
        if 'model_not_found' in raw or 'No available channel' in raw:
            return f"模型不可用：{msg}\n请检查密钥管理中的模型名称是否正确。"
        if 'invalid api key' in raw.lower() or 'authentication' in raw.lower():
            return f"API Key 无效或已过期，请在密钥管理中重新配置。"
        if 'invalid token' in raw.lower():
            return f"供应商返回 Invalid token：{msg}\n如确认 Key 未过期，请检查 Base URL、模型名称是否匹配，或凭 request_id 联系供应商核查。"
        if 'rate limit' in raw.lower():
            return f"请求频率超限，请稍后再试。"
        if 'context_length' in raw.lower() or 'too long' in raw.lower():
            return f"消息过长，请清空对话后重试。"
        return msg  # 返回提取到的 message，已比原始报错简洁
    # 提取状态码
    m2 = _re.search(r'Error code: (\d+)', raw)
    if m2:
        code = m2.group(1)
        return f"API 返回错误（HTTP {code}），请检查配置是否正确。"
    return "请求失败，请检查 API 配置后重试。"

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost", "http://127.0.0.1", "null"],
    allow_origin_regex=r"http://localhost:\d+",
    allow_methods=["*"],
    allow_headers=["*"],
)

MODELS = ["claude-haiku-4-5-20251001", "claude-sonnet-4-6", "claude-opus-4-7"]  # fallback only

class TextBlock(BaseModel):
    type: Literal["text"]
    text: str

class ImageBlock(BaseModel):
    type: Literal["image"]
    source: dict

class DocumentBlock(BaseModel):
    type: Literal["document"]
    source: dict

class Message(BaseModel):
    role: Literal["user", "assistant"]
    content: str | list[TextBlock | ImageBlock | DocumentBlock]

class ChatRequest(BaseModel):
    messages: list[Message]
    model: str = "claude-sonnet-4-6"
    system: str = ""
    config_id: int | None = None
    request_id: int | None = None
    user_token: str | None = None  # 用于记录调用用户

@app.get("/health")
def health():
    return {"status": "ok"}

@app.get("/models")
def models():
    return {"models": MODELS}

def _fetch_models_for_key(api_key: str, base_url: str):
    """共用逻辑:用一个 key 调供应商 /v1/models。返回排序后的列表或 raise。"""
    import httpx
    base_url = (base_url or "").rstrip("/")
    if not base_url.endswith("/v1"):
        base_url = base_url + "/v1"
    try:
        resp = httpx.get(
            f"{base_url}/models",
            headers={"Authorization": f"Bearer {api_key}"},
            timeout=10,
        )
        resp.raise_for_status()
        data = resp.json()
        models = []
        if isinstance(data, dict) and "data" in data:
            models = [m["id"] for m in data["data"] if isinstance(m, dict) and "id" in m]
        elif isinstance(data, dict) and "models" in data:
            raw = data["models"]
            models = [m["id"] if isinstance(m, dict) else str(m) for m in raw]
        elif isinstance(data, list):
            models = [m["id"] if isinstance(m, dict) else str(m) for m in data]
        return sorted(models)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"无法连接供应商:{e}")


@app.get("/admin/configs/{config_id}/fetch-models")
def fetch_models_legacy(config_id: int, x_admin_password: str = Header(default="")):
    """旧路径,config_id 现等于 api_keys.id。"""
    require_admin(x_admin_password)
    with get_db() as conn:
        row = conn.execute(
            "SELECT k.api_key, a.base_url "
            "FROM api_keys k JOIN sub_accounts s ON s.id=k.sub_account_id "
            "JOIN accounts a ON a.id=s.account_id WHERE k.id=?",
            (config_id,),
        ).fetchone()
    if not row:
        raise HTTPException(404, "配置不存在")
    models = _fetch_models_for_key(row["api_key"], row["base_url"])
    return {"models": models, "empty": not models}


@app.get("/admin/accounts/{account_id}/fetch-models")
def fetch_models_for_account(account_id: int, x_token: str = Header(default="")):
    user = get_current_user(x_token)
    with get_db() as conn:
        acc = conn.execute("SELECT * FROM accounts WHERE id=?", (account_id,)).fetchone()
        if not acc:
            raise HTTPException(404, "帐号不存在")
        if user["role"] != "admin" and acc["created_by"] != user["id"]:
            raise HTTPException(403, "无权操作")
        key_row = conn.execute("""
            SELECT k.api_key FROM api_keys k
            JOIN sub_accounts s ON s.id = k.sub_account_id
            WHERE s.account_id = ? AND k.is_active = 1
            ORDER BY k.id DESC LIMIT 1
        """, (account_id,)).fetchone()
    if not key_row:
        raise HTTPException(400, "该帐号下没有可用 key,无法拉取模型")
    models = _fetch_models_for_key(key_row["api_key"], acc["base_url"])
    return {"models": models, "empty": not models}

@app.get("/configs/{config_id}/models")
def config_models(config_id: int):
    """返回指定配置的可用模型列表。"""
    with get_db() as conn:
        row = conn.execute("SELECT models FROM api_configs WHERE id=?", (config_id,)).fetchone()
    if not row or not row["models"]:
        return {"models": []}
    return {"models": [m.strip() for m in row["models"].split(",") if m.strip()]}

@app.get("/active-configs")
def active_configs():
    """返回所有激活的配置列表（仅 id、name、provider，不暴露密钥）。"""
    try:
        with get_db() as conn:
            rows = conn.execute(
                "SELECT id, name, provider, manager FROM api_configs WHERE is_active=1 ORDER BY id DESC"
            ).fetchall()
        return [{"id": r["id"], "name": r["name"], "provider": r["provider"], "manager": r["manager"] or ""} for r in rows]
    except Exception:
        return []

@app.post("/chat")
async def chat(req: ChatRequest):
    if not req.model:
        raise HTTPException(status_code=400, detail="model 不能为空")

    start_time = time.time()
    called_at  = datetime.now(timezone.utc).isoformat()
    effective_key_id: int | None = req.config_id

    def generate():
        nonlocal effective_key_id
        if req.request_id and req.user_token:
            try:
                _user = get_current_user(req.user_token)
                _key, _base_url, _provider, _key_id = get_credentials_by_request(req.request_id, _user["id"])
                effective_key_id = _key_id
            except HTTPException as e:
                yield f"data: {json.dumps({'error': e.detail})}\n\n"
                return
        else:
            _key, _base_url, _provider = get_config_credentials(req.config_id)
        if not _key:
            yield f"data: {json.dumps({'error': '未配置 API Key，请在管理员面板添加配置'})}\n\n"
            return

        # 规范化 base_url
        _normalized_url = (_base_url or '').rstrip('/')

        # 判断是否使用 Anthropic 原生格式
        _use_anthropic = (_provider or '').lower() == 'anthropic' or (
            'anthropic' in _normalized_url and 'api.anthropic.com' in _normalized_url
        )

        input_tokens  = 0
        output_tokens = 0
        success       = 1
        error_msg     = None
        try:
            if _use_anthropic:
                # Anthropic 原生 SDK
                _ak = {"api_key": _key}
                if _normalized_url:
                    _base = _normalized_url[:-3].rstrip('/') if _normalized_url.endswith('/v1') else _normalized_url
                    _ak["base_url"] = _base
                _client = anthropic.Anthropic(**_ak)
                kwargs = dict(model=req.model, max_tokens=4096,
                              messages=[m.model_dump() for m in req.messages])
                if req.system:
                    kwargs["system"] = req.system
                with _client.messages.stream(**kwargs) as stream:
                    for text in stream.text_stream:
                        yield f"data: {json.dumps({'text': text})}\n\n"
                    usage = stream.get_final_message().usage
                    input_tokens  = usage.input_tokens
                    output_tokens = usage.output_tokens
            else:
                # OpenAI 兼容格式（一步API、New API 等代理）
                from openai import OpenAI as _OpenAI
                _base = _normalized_url if _normalized_url else "https://api.openai.com/v1"
                if not _normalized_url.endswith('/v1'):
                    _base = _normalized_url + '/v1' if _normalized_url else "https://api.openai.com/v1"
                _oc = _OpenAI(api_key=_key, base_url=_base)
                # 拼装 messages：Anthropic 格式 → OpenAI 格式
                oai_msgs = []
                if req.system:
                    oai_msgs.append({"role": "system", "content": req.system})
                for m in req.messages:
                    raw = m.model_dump()
                    # content 可能是 JSON 字符串（从 DB 加载的历史消息）
                    if isinstance(raw.get("content"), str):
                        try:
                            import json as _json
                            parsed = _json.loads(raw["content"])
                            if isinstance(parsed, list):
                                raw["content"] = parsed
                        except Exception:
                            pass
                    if isinstance(raw.get("content"), list):
                        oai_content = []
                        for block in raw["content"]:
                            btype = block.get("type", "")
                            if btype == "text":
                                oai_content.append({"type": "text", "text": block.get("text", "")})
                            elif btype == "image":
                                src = block.get("source", {})
                                if src.get("type") == "base64":
                                    url = f"data:{src['media_type']};base64,{src['data']}"
                                    oai_content.append({"type": "image_url", "image_url": {"url": url}})
                        oai_msgs.append({"role": raw["role"], "content": oai_content})
                    else:
                        oai_msgs.append({"role": raw["role"], "content": raw["content"]})
                stream = _oc.chat.completions.create(
                    model=req.model, messages=oai_msgs,
                    max_tokens=4096, stream=True
                )
                for chunk in stream:
                    delta = chunk.choices[0].delta if chunk.choices else None
                    if delta and delta.content:
                        yield f"data: {json.dumps({'text': delta.content})}\n\n"
                    # 从最后一个 chunk 获取用量（部分代理支持）
                    if hasattr(chunk, 'usage') and chunk.usage:
                        input_tokens  = chunk.usage.prompt_tokens or 0
                        output_tokens = chunk.usage.completion_tokens or 0
            yield "data: [DONE]\n\n"
        except Exception as e:
            success   = 0
            error_msg = str(e)
            yield f"data: {json.dumps({'error': _friendly_error(e)})}\n\n"
        finally:
            if effective_key_id is not None:
                duration_ms = int((time.time() - start_time) * 1000)
                try:
                    with get_db() as conn:
                        # 获取该配置的定价
                        cfg = conn.execute(
                            "SELECT price_input, price_output FROM api_configs WHERE id=?", (effective_key_id,)
                        ).fetchone()
                        pi = cfg["price_input"]  if cfg else 0
                        po = cfg["price_output"] if cfg else 0
                        cost = (input_tokens / 1000.0 * pi) + (output_tokens / 1000.0 * po)

                        # 获取调用用户 ID
                        uid = None
                        if req.user_token:
                            row = conn.execute(
                                "SELECT user_id FROM user_tokens WHERE token=?", (req.user_token,)
                            ).fetchone()
                            if row: uid = row["user_id"]

                        conn.execute(
                            "INSERT INTO usage_stats (config_id,user_id,called_at,model,input_tokens,output_tokens,cost,success,duration_ms,error_msg) VALUES (?,?,?,?,?,?,?,?,?,?)",
                            (effective_key_id, uid, called_at, req.model, input_tokens, output_tokens, cost, success, duration_ms, error_msg)
                        )
                        conn.commit()
                except Exception:
                    pass

    return StreamingResponse(generate(), media_type="text/event-stream")

# ── Admin: accounts (Spec 1 三层结构) ─────────────────────

class AccountIn(BaseModel):
    provider: str
    base_url: str
    provider_backend_url: str = ""
    quota_total_path: str = ""
    balance_path: str = ""
    cost_path: str = ""
    quota_total_json_path: str = ""
    balance_json_path: str = ""
    cost_json_path: str = ""
    manager_user_id: int | None = None
    team: str = ""
    models: str = ""
    is_active: int = 1

class SubAccountIn(BaseModel):
    name: str
    description: str = ""

class ApiKeyIn(BaseModel):
    name: str
    api_key: str
    is_active: int = 1
    exhausted: int = 0
    manager_user_id: int | None = None


@app.get("/admin/accounts")
def list_accounts(x_token: str = Header(default="")):
    user = get_current_user(x_token)
    where_sql, where_params = visibility_filter(user)
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT a.*,
                   u_mgr.username AS manager_username,
                   u_cb.username  AS creator_username
            FROM accounts a
            LEFT JOIN users u_mgr ON u_mgr.id = a.manager_user_id
            LEFT JOIN users u_cb  ON u_cb.id  = a.created_by
            WHERE {where_sql}
            ORDER BY a.id DESC
        """, where_params).fetchall()
    return [dict(r) for r in rows]


@app.get("/admin/accounts/options")
def list_account_options(x_token: str = Header(default="")):
    """新增 key 等场景下拉用：返回全部 active accounts（不受 visibility_filter）。"""
    get_current_user(x_token)
    with get_db() as conn:
        rows = conn.execute("""
            SELECT a.id, a.provider, a.base_url, a.manager_user_id,
                   u.username AS manager_username
            FROM accounts a
            LEFT JOIN users u ON u.id = a.manager_user_id
            WHERE a.is_active = 1
            ORDER BY a.provider, a.id
        """).fetchall()
    return [dict(r) for r in rows]


@app.post("/admin/accounts")
def create_account(body: AccountIn, x_token: str = Header(default="")):
    user = get_current_user(x_token)
    now = datetime.now(timezone.utc).isoformat()
    with get_db() as conn:
        dup = conn.execute(
            "SELECT 1 FROM accounts WHERE provider=? AND base_url=? LIMIT 1",
            (body.provider, body.base_url),
        ).fetchone()
        if dup:
            raise HTTPException(409, "已有相同供应商（provider + base_url 完全相同）")
        cur = conn.execute("""
            INSERT INTO accounts
              (provider, base_url, provider_backend_url, quota_total_path,
               balance_path, cost_path,
               quota_total_json_path, balance_json_path, cost_json_path,
               manager_user_id, team, created_by,
               models, is_active, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (body.provider, body.base_url, body.provider_backend_url,
              body.quota_total_path, body.balance_path, body.cost_path,
              body.quota_total_json_path, body.balance_json_path, body.cost_json_path,
              body.manager_user_id, body.team, user["id"], body.models,
              body.is_active, now, now))
        conn.commit()
        row = conn.execute("SELECT * FROM accounts WHERE id=?", (cur.lastrowid,)).fetchone()
    return dict(row)


@app.put("/admin/accounts/{account_id}")
def update_account(account_id: int, body: AccountIn,
                   x_token: str = Header(default="")):
    user = get_current_user(x_token)
    with get_db() as conn:
        acc = conn.execute(
            "SELECT * FROM accounts WHERE id=?", (account_id,)
        ).fetchone()
        if not acc:
            raise HTTPException(status_code=404, detail="不存在")
        require_owner_or_admin(user, acc)
        now = datetime.now(timezone.utc).isoformat()
        conn.execute("""
            UPDATE accounts SET
              provider=?, base_url=?, provider_backend_url=?, quota_total_path=?,
              balance_path=?, cost_path=?, manager_user_id=?, team=?,
              models=?, is_active=?, updated_at=?
            WHERE id=?
        """, (body.provider, body.base_url, body.provider_backend_url,
              body.quota_total_path, body.balance_path, body.cost_path,
              body.manager_user_id, body.team, body.models, body.is_active,
              now, account_id))
        conn.commit()
        row = conn.execute("SELECT * FROM accounts WHERE id=?", (account_id,)).fetchone()
    return dict(row)


@app.delete("/admin/accounts/{account_id}", status_code=204)
def delete_account(account_id: int, x_token: str = Header(default="")):
    user = get_current_user(x_token)
    with get_db() as conn:
        conn.execute("PRAGMA foreign_keys = ON")
        acc = conn.execute(
            "SELECT * FROM accounts WHERE id=?", (account_id,)
        ).fetchone()
        if not acc:
            raise HTTPException(status_code=404, detail="不存在")
        require_owner_or_admin(user, acc)
        used = conn.execute("""
            SELECT 1 FROM usage_stats us
            JOIN api_keys k     ON k.id = us.config_id
            JOIN sub_accounts s ON s.id = k.sub_account_id
            WHERE s.account_id = ?
            LIMIT 1
        """, (account_id,)).fetchone()
        if used:
            raise HTTPException(status_code=400, detail="该帐号下有 API 已被调用,不能删除")
        conn.execute("DELETE FROM accounts WHERE id=?", (account_id,))
        conn.commit()
    return


# ── Admin: sub-accounts ───────────────────────────────────

@app.get("/admin/accounts/{account_id}/sub-accounts")
def list_sub_accounts_new(account_id: int, x_token: str = Header(default="")):
    get_current_user(x_token)
    with get_db() as conn:
        acc = conn.execute("SELECT * FROM accounts WHERE id=?", (account_id,)).fetchone()
        if not acc:
            raise HTTPException(404, "不存在")
        rows = conn.execute(
            "SELECT * FROM sub_accounts WHERE account_id=? ORDER BY id DESC",
            (account_id,),
        ).fetchall()
    return [dict(r) for r in rows]


@app.post("/admin/accounts/{account_id}/sub-accounts")
def create_sub_account_new(account_id: int, body: SubAccountIn,
                            x_token: str = Header(default="")):
    user = get_current_user(x_token)
    with get_db() as conn:
        acc = conn.execute("SELECT * FROM accounts WHERE id=?", (account_id,)).fetchone()
        if not acc:
            raise HTTPException(404, "不存在")
        require_owner_or_admin(user, acc)
        now = datetime.now(timezone.utc).isoformat()
        cur = conn.execute(
            "INSERT INTO sub_accounts (account_id, name, description, created_at) "
            "VALUES (?, ?, ?, ?)",
            (account_id, body.name, body.description, now),
        )
        conn.commit()
        row = conn.execute("SELECT * FROM sub_accounts WHERE id=?", (cur.lastrowid,)).fetchone()
    return dict(row)


def _get_sub_account_or_404(conn, sub_id: int):
    row = conn.execute("""
        SELECT s.*, a.created_by AS account_created_by,
               a.manager_user_id AS account_manager,
               a.provider AS account_provider
        FROM sub_accounts s JOIN accounts a ON a.id = s.account_id
        WHERE s.id = ?
    """, (sub_id,)).fetchone()
    if not row:
        raise HTTPException(404, "子帐号不存在")
    return row


@app.put("/admin/sub-accounts/{sub_id}")
def update_sub_account_new(sub_id: int, body: SubAccountIn,
                            x_token: str = Header(default="")):
    user = get_current_user(x_token)
    with get_db() as conn:
        sub = _get_sub_account_or_404(conn, sub_id)
        if user["role"] != "admin" and sub["account_created_by"] != user["id"]:
            raise HTTPException(403, "无权操作")
        conn.execute(
            "UPDATE sub_accounts SET name=?, description=? WHERE id=?",
            (body.name, body.description, sub_id),
        )
        conn.commit()
        row = conn.execute("SELECT * FROM sub_accounts WHERE id=?", (sub_id,)).fetchone()
    return dict(row)


@app.delete("/admin/sub-accounts/{sub_id}", status_code=204)
def delete_sub_account_new(sub_id: int, x_token: str = Header(default="")):
    user = get_current_user(x_token)
    with get_db() as conn:
        sub = _get_sub_account_or_404(conn, sub_id)
        if user["role"] != "admin" and sub["account_created_by"] != user["id"]:
            raise HTTPException(403, "无权操作")
        has_keys = conn.execute(
            "SELECT 1 FROM api_keys WHERE sub_account_id=? LIMIT 1", (sub_id,)
        ).fetchone()
        if has_keys:
            raise HTTPException(400, "子帐号下有 API key,不能删除")
        conn.execute("DELETE FROM sub_accounts WHERE id=?", (sub_id,))
        conn.commit()
    return


# ── Admin: api_keys ───────────────────────────────────────

@app.get("/admin/sub-accounts/{sub_id}/api-keys")
def list_api_keys(sub_id: int, x_token: str = Header(default="")):
    user = get_current_user(x_token)
    with get_db() as conn:
        _get_sub_account_or_404(conn, sub_id)
        if user["role"] == "admin":
            key_filter = ""
            params = (sub_id,)
        else:
            key_filter = "AND (k.created_by = ? OR k.manager_user_id = ?)"
            params = (sub_id, user["id"], user["id"])
        rows = conn.execute(f"""
            SELECT k.*, u.username AS manager_username,
                   uc.username AS creator_username,
                   (SELECT COUNT(*) FROM usage_stats WHERE config_id = k.id) AS usage_count,
                   (SELECT GROUP_CONCAT(DISTINCT au.username)
                      FROM api_requests ar
                      JOIN users au ON au.id = ar.user_id
                      WHERE ar.api_key_id = k.id AND ar.status = 'approved') AS users_username
            FROM api_keys k
            LEFT JOIN users u  ON u.id  = k.manager_user_id
            LEFT JOIN users uc ON uc.id = k.created_by
            WHERE k.sub_account_id=? {key_filter}
            ORDER BY k.id DESC
        """, params).fetchall()
    return [dict(r) for r in rows]


@app.post("/admin/sub-accounts/{sub_id}/api-keys")
def create_api_key(sub_id: int, body: ApiKeyIn,
                   x_token: str = Header(default="")):
    user = get_current_user(x_token)
    with get_db() as conn:
        sub = _get_sub_account_or_404(conn, sub_id)
        provider = sub["account_provider"]
        dup = conn.execute("""
            SELECT 1 FROM api_keys k
            JOIN sub_accounts s ON s.id = k.sub_account_id
            JOIN accounts a     ON a.id = s.account_id
            WHERE a.provider = ? AND k.api_key = ?
            LIMIT 1
        """, (provider, body.api_key)).fetchone()
        if dup:
            raise HTTPException(409, "该 provider 下已存在相同 key 字符串")
        now = datetime.now(timezone.utc).isoformat()
        cur = conn.execute(
            "INSERT INTO api_keys (sub_account_id, name, api_key, is_active, "
            "exhausted, manager_user_id, created_by, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (sub_id, body.name, body.api_key, body.is_active,
             body.exhausted, body.manager_user_id, user["id"], now),
        )
        conn.commit()
        row = conn.execute("""
            SELECT k.*, u.username AS manager_username,
                   uc.username AS creator_username
            FROM api_keys k
            LEFT JOIN users u  ON u.id  = k.manager_user_id
            LEFT JOIN users uc ON uc.id = k.created_by
            WHERE k.id=?
        """, (cur.lastrowid,)).fetchone()
    return dict(row)


def _get_api_key_or_404(conn, key_id: int):
    row = conn.execute("""
        SELECT k.*
        FROM api_keys k
        JOIN sub_accounts s ON s.id = k.sub_account_id
        JOIN accounts a     ON a.id = s.account_id
        WHERE k.id = ?
    """, (key_id,)).fetchone()
    if not row:
        raise HTTPException(404, "key 不存在")
    return row


@app.put("/admin/api-keys/{key_id}")
def update_api_key(key_id: int, body: ApiKeyIn,
                   x_token: str = Header(default="")):
    user = get_current_user(x_token)
    with get_db() as conn:
        key = _get_api_key_or_404(conn, key_id)
        if user["role"] != "admin" and key["created_by"] != user["id"]:
            raise HTTPException(403, "无权操作")
        conn.execute(
            "UPDATE api_keys SET name=?, api_key=?, is_active=?, exhausted=?, manager_user_id=? WHERE id=?",
            (body.name, body.api_key, body.is_active, body.exhausted, body.manager_user_id, key_id),
        )
        conn.commit()
        row = conn.execute("""
            SELECT k.*, u.username AS manager_username,
                   uc.username AS creator_username
            FROM api_keys k
            LEFT JOIN users u  ON u.id  = k.manager_user_id
            LEFT JOIN users uc ON uc.id = k.created_by
            WHERE k.id=?
        """, (key_id,)).fetchone()
    return dict(row)


@app.delete("/admin/api-keys/{key_id}", status_code=204)
def delete_api_key(key_id: int, x_token: str = Header(default="")):
    user = get_current_user(x_token)
    with get_db() as conn:
        key = _get_api_key_or_404(conn, key_id)
        if user["role"] != "admin" and key["created_by"] != user["id"]:
            raise HTTPException(403, "无权操作")
        used = conn.execute(
            "SELECT 1 FROM usage_stats WHERE config_id=? LIMIT 1", (key_id,)
        ).fetchone()
        if used:
            raise HTTPException(400, "已被调用,不能删除")
        conn.execute("DELETE FROM api_keys WHERE id=?", (key_id,))
        conn.commit()
    return


# ── Admin: providers / teams (Spec 1) ─────────────────────

@app.get("/admin/providers")
def list_providers_spec1(x_token: str = Header(default="")):
    user = get_current_user(x_token)
    where_sql, where_params = visibility_filter(user)
    with get_db() as conn:
        rows = conn.execute(
            f"SELECT DISTINCT a.provider FROM accounts a "
            f"WHERE {where_sql} AND a.provider <> ''",
            where_params,
        ).fetchall()
    return [r["provider"] for r in rows]


@app.get("/admin/teams")
def list_teams(x_token: str = Header(default="")):
    user = get_current_user(x_token)
    where_sql, where_params = visibility_filter(user)
    with get_db() as conn:
        rows = conn.execute(
            f"SELECT DISTINCT a.team FROM accounts a "
            f"WHERE {where_sql} AND a.team <> ''",
            where_params,
        ).fetchall()
    return [r["team"] for r in rows]


# ── Spec 2: quota-all + Excel import ────────────────────

import httpx as _httpx
from providers import extract_json_value


async def _fetch_one_key(r: dict) -> tuple:
    base = (r["provider_backend_url"] or r["base_url"] or "").rstrip("/")
    headers = {"Authorization": f"Bearer {r['api_key']}"}

    async def _get(path):
        if not path:
            return None
        async with _httpx.AsyncClient(timeout=10) as cli:
            resp = await cli.get(base + path, headers=headers)
            resp.raise_for_status()
            return resp.json()

    raw = await asyncio.gather(
        _get(r["quota_total_path"]),
        _get(r["balance_path"]),
        _get(r["cost_path"]),
        return_exceptions=True,
    )

    def _coerce(val, json_path):
        """有 json_path 走 extract_json_value;没有则尝试把原始响应直接当数字。"""
        if isinstance(val, Exception) or val is None:
            return None
        if json_path:
            return extract_json_value(val, json_path)
        # 裸响应直接转 float（适合返回 42.5 / "42.5" 的简单计费接口）
        if isinstance(val, (int, float)):
            return float(val)
        if isinstance(val, str):
            try: return float(val.strip())
            except ValueError: return None
        return None

    total   = _coerce(raw[0], r["quota_total_json_path"])
    balance = _coerce(raw[1], r["balance_json_path"])
    used    = _coerce(raw[2], r["cost_json_path"])
    return total, balance, used


@app.get("/admin/accounts/quota-all")
async def quota_all(x_token: str = Header(default="")):
    user = get_current_user(x_token)
    where_sql, where_params = visibility_filter(user)
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT k.*, a.provider, a.base_url, a.provider_backend_url,
                   a.quota_total_path, a.balance_path, a.cost_path,
                   a.quota_total_json_path, a.balance_json_path, a.cost_json_path
            FROM api_keys k
            JOIN sub_accounts s ON s.id = k.sub_account_id
            JOIN accounts a     ON a.id = s.account_id
            WHERE {where_sql}
        """, where_params).fetchall()
        rows = [dict(r) for r in rows]

    results = {}
    pending = []
    for r in rows:
        if r["exhausted"]:
            results[str(r["id"])] = {
                "total": r.get("last_total"),
                "balance": 0,
                "used": r.get("last_used"),
                "exhausted": True,
                "from_cache": True,
                "cached_at": r.get("last_quota_at"),
            }
        else:
            pending.append(r)

    tasks = [asyncio.create_task(_fetch_one_key(r)) for r in pending]
    fetched = await asyncio.gather(*tasks, return_exceptions=True)
    now_iso = datetime.now(timezone.utc).isoformat()

    for r, result in zip(pending, fetched):
        kid = str(r["id"])
        if isinstance(result, Exception):
            results[kid] = {"error": str(result)}
            continue
        total, balance, used = result
        partial = (total is None) or (balance is None) or (used is None)

        new_zero_count = r.get("zero_count") or 0
        new_last_zero_at = r.get("last_zero_at")
        new_exhausted = 0

        if balance is not None and balance == 0:
            now_dt = datetime.now(timezone.utc)
            should_count = True
            if new_last_zero_at:
                try:
                    last = datetime.fromisoformat(new_last_zero_at)
                    if (now_dt - last).total_seconds() < 12 * 3600:
                        should_count = False
                except Exception:
                    pass
            if should_count:
                new_zero_count += 1
                new_last_zero_at = now_iso
            if new_zero_count >= 3:
                new_exhausted = 1
        elif balance is not None and balance > 0:
            new_zero_count = 0
            new_last_zero_at = None

        results[kid] = {
            "total": total, "balance": balance, "used": used,
            "exhausted": bool(new_exhausted), "from_cache": False,
            "partial": partial,
        }
        with get_db() as conn:
            conn.execute("""
                UPDATE api_keys
                SET last_total=?, last_balance=?, last_used=?, last_quota_at=?,
                    zero_count=?, last_zero_at=?, exhausted=?
                WHERE id=?
            """, (total, balance, used, now_iso,
                  new_zero_count, new_last_zero_at, new_exhausted, r["id"]))
            conn.commit()

    return {"fetched_at": now_iso, "results": results}


@app.post("/admin/sub-accounts/{sub_id}/api-keys/import-excel")
def import_keys_xlsx(sub_id: int, file: UploadFile = File(...),
                     manager_user_id: int | None = Form(default=None),
                     x_token: str = Header(default="")):
    user = get_current_user(x_token)
    with get_db() as conn:
        _get_sub_account_or_404(conn, sub_id)

    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(400, "仅支持 .xlsx 格式")

    from openpyxl import load_workbook
    from io import BytesIO
    content = file.file.read()
    wb = load_workbook(BytesIO(content), read_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        raise HTTPException(400, "Excel 为空")

    header = [str(c or "").strip().lower() for c in all_rows[0]]
    name_col = next((i for i, h in enumerate(header) if "名称" in h or "name" in h), -1)
    key_col = next((i for i, h in enumerate(header) if "key" in h), -1)
    if name_col < 0 or key_col < 0:
        raise HTTPException(400, "未找到 API名称 / API key 列")

    imported, skipped, errors = 0, [], []
    for idx, row in enumerate(all_rows[1:], start=2):
        name = str(row[name_col] or "").strip() if name_col < len(row) else ""
        key = str(row[key_col] or "").strip() if key_col < len(row) else ""
        if not name or not key:
            continue  # 空行静默跳过，不计入失败
        try:
            create_api_key(sub_id, ApiKeyIn(name=name, api_key=key, manager_user_id=manager_user_id), x_token)
            imported += 1
        except HTTPException as e:
            if e.status_code == 409:
                skipped.append({"row": idx, "name": name})
            else:
                errors.append({"row": idx, "reason": str(e.detail)})

    return {"imported": imported, "skipped_duplicates": skipped,
            "errors": errors, "total_rows": len(all_rows) - 1}


@app.get("/admin/api-keys/template.xlsx")
def download_template(x_token: str = Header(default="")):
    get_current_user(x_token)
    from openpyxl import Workbook
    from io import BytesIO
    wb = Workbook()
    ws = wb.active
    ws.title = "API keys"
    ws.append(["API名称", "API key"])
    ws.append(["示例-主key", "sk-xxxxxxxxxxxxxxxx"])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="api_keys_template.xlsx"'},
    )


# ── Admin: configs ────────────────────────────────────────

class ConfigIn(BaseModel):
    name: str
    base_url: str
    api_key: str
    provider: str
    models: str = ""
    price_input: float = 0   # 元/千 input token
    price_output: float = 0  # 元/千 output token
    is_active: int = 1
    manager: str = ""

@app.post("/admin/login")
def admin_login(x_admin_password: str = Header(default="")):
    return {"ok": x_admin_password == ADMIN_PASSWORD}

@app.get("/admin/configs")
def list_configs(x_admin_password: str = Header(default="")):
    require_admin(x_admin_password)
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM api_configs ORDER BY id DESC").fetchall()
    result = []
    for r in rows:
        d = dict(r)
        d["api_key"] = "****" + d["api_key"][-4:] if len(d["api_key"]) >= 4 else "****"
        result.append(d)
    return result

# Spec 1: POST/PUT/DELETE /admin/configs 已移除,改用 /admin/accounts + /admin/api-keys


# ── Admin: stats ───────────────────────────────────────────

@app.get("/admin/stats")
def list_stats(days: int | None = None, x_admin_password: str = Header(default="")):
    require_admin(x_admin_password)
    date_clause = f"AND called_at >= date('now','-{int(days)} days')" if days else ""
    with get_db() as conn:
        configs = conn.execute("SELECT id, name, provider, created_at FROM api_configs").fetchall()
        result = []
        for cfg in configs:
            cid = cfg["id"]
            row = conn.execute(f"""
                SELECT COUNT(*) as total,
                       SUM(success) as ok,
                       SUM(input_tokens+output_tokens) as tokens,
                       SUM(input_tokens) as in_tokens,
                       SUM(output_tokens) as out_tokens,
                       ROUND(SUM(cost),4) as total_cost,
                       AVG(duration_ms) as avg_ms,
                       MAX(called_at) as last_used
                FROM usage_stats WHERE config_id=? {date_clause}
            """, (cid,)).fetchone()
            total = row["total"] or 0
            ok    = row["ok"] or 0
            result.append({
                "config_id":       cid,
                "name":            cfg["name"],
                "provider":        cfg["provider"],
                "created_at":      cfg["created_at"],
                "total_calls":     total,
                "success_rate":    round(ok / total * 100, 1) if total else 0,
                "input_tokens":    row["in_tokens"] or 0,
                "output_tokens":   row["out_tokens"] or 0,
                "total_tokens":    row["tokens"] or 0,
                "total_cost":      row["total_cost"] or 0,
                "avg_duration_ms": round(row["avg_ms"] or 0, 1),
                "last_used":       row["last_used"] or "—",
            })
    return result

@app.get("/admin/stats/by-user")
def stats_by_user(x_admin_password: str = Header(default="")):
    """按用户分组的调用统计（token + 费用）。"""
    require_admin(x_admin_password)
    with get_db() as conn:
        rows = conn.execute("""
            SELECT
                COALESCE(u.username, '匿名') as username,
                COUNT(*) as total_calls,
                SUM(s.input_tokens)  as input_tokens,
                SUM(s.output_tokens) as output_tokens,
                SUM(s.input_tokens + s.output_tokens) as total_tokens,
                ROUND(SUM(s.cost), 4) as total_cost,
                SUM(s.success) as success_count,
                MAX(s.called_at) as last_used
            FROM usage_stats s
            LEFT JOIN users u ON u.id = s.user_id
            GROUP BY s.user_id
            ORDER BY total_cost DESC
        """).fetchall()
    return [dict(r) for r in rows]

@app.get("/admin/stats/{config_id}/daily")
def daily_stats(config_id: int, days: int = 7, x_admin_password: str = Header(default="")):
    require_admin(x_admin_password)
    n = max(1, int(days))
    with get_db() as conn:
        rows = conn.execute("""
            SELECT substr(called_at,1,10) as day, COUNT(*) as cnt
            FROM usage_stats
            WHERE config_id=?
              AND called_at >= date('now',? || ' days')
            GROUP BY day ORDER BY day
        """, (config_id, f"-{n}")).fetchall()
    return [{"day": r["day"], "count": r["cnt"]} for r in rows]


# ── Platform 看板（admin） ──────────────────────────────

@app.get("/admin/stats/platform/overview")
def platform_overview(days: int = 1, x_admin_password: str = Header(default="")):
    """核心指标。days=1 今日,days=N 近 N 天,days=0 总量(全平台累计)。"""
    require_admin(x_admin_password)
    n = max(0, min(int(days), 36500))
    if n == 0:
        date_filter = "1=1"
    elif n == 1:
        date_filter = "date(called_at) = date('now','localtime')"
    else:
        date_filter = f"called_at >= date('now','-{n-1} days')"
    # 上面是给 usage_stats 用的;给 join 后表用,前缀 u.
    u_filter = date_filter.replace("called_at", "u.called_at")

    with get_db() as conn:
        row = conn.execute(f"""
            SELECT COUNT(*) AS total,
                   SUM(success) AS ok,
                   COALESCE(SUM(input_tokens + output_tokens), 0) AS tokens,
                   COALESCE(SUM(cost), 0) AS cost,
                   COALESCE(AVG(duration_ms), 0) AS avg_ms,
                   COUNT(DISTINCT user_id) AS active_users
            FROM usage_stats
            WHERE {date_filter}
        """).fetchone()
        # 活跃项目:期间内有调用的不同 project_name 列表
        proj_rows = conn.execute(f"""
            SELECT DISTINCT COALESCE(r.project_name,'未知项目') AS project
            FROM usage_stats u
            JOIN api_requests r ON r.user_id = u.user_id AND r.api_key_id = u.config_id
            WHERE {u_filter} AND r.status = 'approved'
            ORDER BY project
        """).fetchall()
    total = row["total"] or 0
    ok    = row["ok"] or 0
    projects = [r["project"] for r in proj_rows]
    return {
        "total_calls":  total,
        "total_tokens": row["tokens"] or 0,
        "total_cost":   round(row["cost"] or 0, 4),
        "active_projects": projects,        # 项目名称数组
        "active_projects_count": len(projects),
        "active_users": row["active_users"] or 0,
        "success_rate": round(ok / total * 100, 1) if total else 0,
        "avg_duration_ms": int(row["avg_ms"] or 0),
        "days": n,
    }


@app.get("/admin/stats/platform/trends")
def platform_trends(days: int = 7, top: int = 5, x_admin_password: str = Header(default="")):
    """请求/成本/Top N 模型每日趋势;days=0 总量;top 默认 5(展开时传 10)。"""
    require_admin(x_admin_password)
    n = max(0, min(int(days), 36500))
    top_n = max(1, min(int(top), 20))
    if n == 0: date_filter = "1=1"
    elif n == 1: date_filter = "date(called_at) = date('now','localtime')"
    else: date_filter = f"called_at >= date('now','-{n-1} days')"
    with get_db() as conn:
        daily = conn.execute(f"""
            SELECT date(called_at) AS day,
                   COUNT(*) AS calls,
                   COALESCE(SUM(cost), 0) AS cost
            FROM usage_stats
            WHERE {date_filter}
            GROUP BY day ORDER BY day
        """).fetchall()
        top_models = [r["model"] for r in conn.execute(f"""
            SELECT COALESCE(model,'—') AS model, COUNT(*) AS c
            FROM usage_stats
            WHERE {date_filter}
            GROUP BY model ORDER BY c DESC LIMIT {top_n}
        """).fetchall()]
        model_daily = []
        if top_models:
            ph = ",".join("?" * len(top_models))
            model_daily = [dict(r) for r in conn.execute(f"""
                SELECT date(called_at) AS day, COALESCE(model,'—') AS model, COUNT(*) AS calls
                FROM usage_stats
                WHERE {date_filter}
                  AND COALESCE(model,'—') IN ({ph})
                GROUP BY day, model ORDER BY day
            """, top_models).fetchall()]
    return {
        "days":         n,
        "daily":        [dict(r) for r in daily],
        "top_models":   top_models,
        "model_daily":  model_daily,
    }


@app.get("/admin/stats/platform/ranking")
def platform_ranking(days: int = 7, limit: int = 10, x_admin_password: str = Header(default="")):
    """项目 + 用户 调用排行;默认 Top10,limit 可放大(上限 1000)用于展开查看全部;days=0 为总量。"""
    require_admin(x_admin_password)
    n = max(0, min(int(days), 36500))
    lim = max(1, min(int(limit), 1000))
    if n == 0: u_filter = "1=1"
    elif n == 1: u_filter = "date(u.called_at) = date('now','localtime')"
    else: u_filter = f"u.called_at >= date('now','-{n-1} days')"
    with get_db() as conn:
        projects = [dict(r) for r in conn.execute(f"""
            SELECT COALESCE(r.project_name,'未知项目') AS project,
                   COUNT(*) AS calls,
                   COALESCE(SUM(u.input_tokens + u.output_tokens),0) AS tokens,
                   COALESCE(SUM(u.cost),0) AS cost
            FROM usage_stats u
            LEFT JOIN api_requests r ON r.user_id = u.user_id AND r.api_key_id = u.config_id
            WHERE {u_filter}
            GROUP BY project
            ORDER BY calls DESC LIMIT {lim}
        """).fetchall()]
        users = [dict(r) for r in conn.execute(f"""
            SELECT COALESCE(usr.username,'未知') AS username,
                   COUNT(*) AS calls,
                   COALESCE(SUM(u.input_tokens + u.output_tokens),0) AS tokens,
                   COALESCE(SUM(u.cost),0) AS cost
            FROM usage_stats u
            LEFT JOIN users usr ON usr.id = u.user_id
            WHERE {u_filter}
            GROUP BY u.user_id
            ORDER BY calls DESC LIMIT {lim}
        """).fetchall()]
    for r in projects + users:
        r["cost"] = round(r["cost"], 4)
    return {"projects": projects, "users": users, "days": n}


@app.get("/admin/stats/platform/models")
def platform_models(days: int = 7, x_admin_password: str = Header(default="")):
    """模型治理：成本/成功率/延迟 + 热门 Top5;days=0 总量。"""
    require_admin(x_admin_password)
    n = max(0, min(int(days), 36500))
    if n == 0: date_filter = "1=1"
    elif n == 1: date_filter = "date(called_at) = date('now','localtime')"
    else: date_filter = f"called_at >= date('now','-{n-1} days')"
    with get_db() as conn:
        rows = [dict(r) for r in conn.execute(f"""
            SELECT COALESCE(model,'—') AS model,
                   COUNT(*) AS calls,
                   SUM(success) AS ok,
                   COALESCE(SUM(cost),0) AS cost,
                   COALESCE(AVG(duration_ms),0) AS avg_ms,
                   COALESCE(SUM(input_tokens+output_tokens),0) AS tokens
            FROM usage_stats
            WHERE {date_filter}
            GROUP BY model
            ORDER BY calls DESC
        """).fetchall()]
    for r in rows:
        c = r["calls"] or 0
        r["success_rate"] = round((r["ok"] or 0) / c * 100, 1) if c else 0
        r["avg_ms"] = int(r["avg_ms"] or 0)
        r["cost"]   = round(r["cost"], 4)
        r.pop("ok", None)
    return {"models": rows, "days": n}


@app.get("/admin/stats/platform/anomalies")
def platform_anomalies(x_admin_password: str = Header(default="")):
    """失败率 / 超时（>10s） / 今日 vs 昨日同比 +50% 标记。"""
    require_admin(x_admin_password)
    with get_db() as conn:
        today = conn.execute("""
            SELECT COUNT(*) AS total, SUM(success) AS ok,
                   SUM(CASE WHEN duration_ms > 10000 THEN 1 ELSE 0 END) AS slow,
                   SUM(CASE WHEN success = 0 THEN 1 ELSE 0 END) AS fails
            FROM usage_stats
            WHERE date(called_at) = date('now','localtime')
        """).fetchone()
        yest = conn.execute("""
            SELECT COUNT(*) AS total,
                   SUM(CASE WHEN success = 0 THEN 1 ELSE 0 END) AS fails
            FROM usage_stats
            WHERE date(called_at) = date('now','localtime','-1 day')
        """).fetchone()
        # 今日按小时
        hourly = [dict(r) for r in conn.execute("""
            SELECT strftime('%H', called_at) AS hour,
                   COUNT(*) AS calls,
                   SUM(CASE WHEN success=0 THEN 1 ELSE 0 END) AS fails
            FROM usage_stats
            WHERE date(called_at) = date('now','localtime')
            GROUP BY hour ORDER BY hour
        """).fetchall()]
        # 最近 10 条失败
        recent_fails = [dict(r) for r in conn.execute("""
            SELECT u.called_at, u.model, u.duration_ms, u.error_msg,
                   COALESCE(usr.username,'—') AS username
            FROM usage_stats u
            LEFT JOIN users usr ON usr.id = u.user_id
            WHERE u.success = 0
            ORDER BY u.id DESC LIMIT 10
        """).fetchall()]
    t_total = today["total"] or 0
    y_total = yest["total"] or 0
    t_fails = today["fails"] or 0
    growth_pct = round((t_total - y_total) / y_total * 100, 1) if y_total else None
    return {
        "today_total":   t_total,
        "today_fails":   t_fails,
        "today_slow":    today["slow"] or 0,
        "fail_rate":     round(t_fails / t_total * 100, 1) if t_total else 0,
        "yesterday_total": y_total,
        "growth_pct":    growth_pct,
        "is_anomaly":    growth_pct is not None and growth_pct >= 50,
        "hourly":        hourly,
        "recent_fails":  recent_fails,
    }


# ── Me: 个人看板 stats ─────────────────────────────────────

@app.get("/me/stats/today")
def me_stats_today(x_token: str = Header(default="")):
    """当前用户当日：总请求 / 总 Token / 总费用 / 成功率。"""
    user = get_current_user(x_token)
    with get_db() as conn:
        row = conn.execute("""
            SELECT COUNT(*) AS total,
                   SUM(success) AS ok,
                   COALESCE(SUM(input_tokens + output_tokens), 0) AS tokens,
                   COALESCE(SUM(cost), 0) AS cost
            FROM usage_stats
            WHERE user_id = ?
              AND date(called_at) = date('now', 'localtime')
        """, (user["id"],)).fetchone()
    total = row["total"] or 0
    ok    = row["ok"] or 0
    return {
        "total_calls":  total,
        "total_tokens": row["tokens"] or 0,
        "total_cost":   round(row["cost"] or 0, 4),
        "success_rate": round(ok / total * 100, 1) if total else 0,
    }


@app.get("/me/stats/by-model")
def me_stats_by_model(days: int | None = None, x_token: str = Header(default="")):
    """当前用户按模型分组：调用量、Token 数、费用、占比。"""
    user = get_current_user(x_token)
    date_clause = f"AND called_at >= date('now','-{int(days)} days')" if days else ""
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT COALESCE(model, '—') AS model,
                   COUNT(*) AS calls,
                   COALESCE(SUM(input_tokens), 0)  AS in_tokens,
                   COALESCE(SUM(output_tokens), 0) AS out_tokens,
                   COALESCE(SUM(input_tokens + output_tokens), 0) AS tokens,
                   COALESCE(SUM(cost), 0) AS cost
            FROM usage_stats
            WHERE user_id = ? {date_clause}
            GROUP BY model
            ORDER BY calls DESC
        """, (user["id"],)).fetchall()
    rows = [dict(r) for r in rows]
    total_calls = sum(r["calls"] for r in rows) or 0
    for r in rows:
        r["percent"]   = round(r["calls"] / total_calls * 100, 1) if total_calls else 0
        r["cost"]      = round(r["cost"], 4)
    return rows


@app.get("/me/stats/keys-balance")
def me_stats_keys_balance(x_token: str = Header(default="")):
    """当前用户已批准的 API key 余额视图：总额、已花费、余额。"""
    user = get_current_user(x_token)
    with get_db() as conn:
        rows = conn.execute("""
            SELECT r.id          AS request_id,
                   r.project_name,
                   a.provider,
                   sa.name        AS sub_account_name,
                   k.id           AS api_key_id,
                   k.name         AS api_name,
                   k.last_total   AS total,
                   k.last_balance AS balance,
                   k.last_used    AS used,
                   k.last_quota_at,
                   k.exhausted
            FROM api_requests r
            JOIN api_keys k     ON k.id = r.api_key_id
            JOIN sub_accounts sa ON sa.id = k.sub_account_id
            JOIN accounts a     ON a.id = sa.account_id
            WHERE r.user_id = ? AND r.status = 'approved'
            ORDER BY r.created_at DESC
        """, (user["id"],)).fetchall()
    return [dict(r) for r in rows]


# ── Auth ──────────────────────────────────────────────────

class RegisterIn(BaseModel):
    username: str
    password: str

class LoginIn(BaseModel):
    username: str
    password: str

def get_current_user(x_token: str = Header(default="")):
    if not x_token:
        raise HTTPException(status_code=401, detail="未登录")
    with get_db() as conn:
        row = conn.execute(
            "SELECT u.* FROM users u JOIN user_tokens t ON t.user_id=u.id WHERE t.token=?", (x_token,)
        ).fetchone()
    if not row:
        raise HTTPException(status_code=401, detail="登录已过期")
    return dict(row)


def visibility_filter(user: dict) -> tuple[str, tuple]:
    """密钥管理列表的可见性：admin 看全部；其他人看自己创建的、或自己管理其中 key 的 account。"""
    if user["role"] == "admin":
        return "1=1", ()
    return ("(a.created_by = ? OR a.manager_user_id = ? OR a.id IN "
            "(SELECT s.account_id FROM sub_accounts s JOIN api_keys k ON k.sub_account_id = s.id "
            "WHERE k.manager_user_id = ?))",
            (user["id"], user["id"], user["id"]))


def require_owner_or_admin(user: dict, account: dict) -> None:
    """写权限：admin role、创建人或该帐号的管理员（manager_user_id）。"""
    if user["role"] == "admin":
        return
    if account["created_by"] == user["id"]:
        return
    if account.get("manager_user_id") == user["id"]:
        return
    raise HTTPException(status_code=403, detail="无权操作此帐号")


@app.post("/auth/register")
def register(body: RegisterIn):
    if not body.username.strip() or not body.password.strip():
        raise HTTPException(status_code=400, detail="用户名和密码不能为空")
    now = datetime.now(timezone.utc).isoformat()
    try:
        with get_db() as conn:
            conn.execute(
                "INSERT INTO users (username, password, created_at) VALUES (?,?,?)",
                (body.username.strip(), hash_password(body.password), now)
            )
            conn.commit()
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=400, detail="用户名已存在")
    return {"ok": True}

@app.post("/auth/login")
def user_login(body: LoginIn):
    with get_db() as conn:
        user = conn.execute(
            "SELECT * FROM users WHERE username=? AND password=?",
            (body.username.strip(), hash_password(body.password))
        ).fetchone()
    if not user:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = secrets.token_hex(32)
    now = datetime.now(timezone.utc).isoformat()
    with get_db() as conn:
        conn.execute("INSERT INTO user_tokens (token, user_id, created_at) VALUES (?,?,?)",
                     (token, user["id"], now))
        conn.commit()
    return {"token": token, "user": {"id": user["id"], "username": user["username"], "role": user["role"]}}

@app.post("/auth/logout")
def user_logout(x_token: str = Header(default="")):
    with get_db() as conn:
        conn.execute("DELETE FROM user_tokens WHERE token=?", (x_token,))
        conn.commit()
    return {"ok": True}

@app.get("/auth/me")
def me(x_token: str = Header(default="")):
    return get_current_user(x_token)

# ── API 申请 ───────────────────────────────────────────────

class RequestIn(BaseModel):
    account_id: int | None = None
    project_name: str
    lead: str
    budget: str
    purpose: str = ""           # 需求详情
    reviewer_id: int | None = None   # 审批人 user_id（必须为 account 的管理员）
    dept: str = ""
    # 向后兼容旧字段
    config_id: int | None = None
    sub_accounts: str = ""
    cc_person: str = ""

@app.get("/api-requests/cascade-options")
def cascade_options():
    """返回级联选项：(account × key级管理员) 笛卡尔对，前端先选管理员再过滤 account。
    每个 (account, manager) 组合一行，去重。"""
    with get_db() as conn:
        rows = conn.execute("""
            SELECT DISTINCT a.id, a.provider, a.team,
                   u.username AS manager_username, u.id AS manager_user_id
            FROM accounts a
            JOIN sub_accounts s ON s.account_id = a.id
            JOIN api_keys k    ON k.sub_account_id = s.id
            JOIN users u       ON u.id = k.manager_user_id
            WHERE a.is_active = 1 AND k.is_active = 1
            ORDER BY a.provider, u.username
        """).fetchall()
    return [dict(r) for r in rows]

@app.post("/api-requests")
def create_request(body: RequestIn, x_token: str = Header(default="")):
    user = get_current_user(x_token)
    if not body.account_id:
        raise HTTPException(400, "请选择供应商")
    if not body.reviewer_id:
        raise HTTPException(400, "请选择审批人")
    now = datetime.now(timezone.utc).isoformat()
    with get_db() as conn:
        # 校验 reviewer_id 是否为该 account 下任意 api_keys 的管理员（key 级）
        owns = conn.execute("""
            SELECT 1 FROM api_keys k
            JOIN sub_accounts s ON s.id = k.sub_account_id
            WHERE s.account_id = ? AND k.manager_user_id = ? AND k.is_active = 1
            LIMIT 1
        """, (body.account_id, body.reviewer_id)).fetchone()
        if not owns:
            raise HTTPException(400, "该成员名下没有此 API")
        cur = conn.execute(
            "INSERT INTO api_requests (user_id,config_id,account_id,project_name,purpose,lead,budget,dept,sub_accounts,cc_person,reviewer_id,created_at,updated_at) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (user["id"], body.config_id, body.account_id, body.project_name,
             body.purpose, body.lead, body.budget, body.dept,
             body.sub_accounts, body.cc_person, body.reviewer_id, now, now)
        )
        conn.commit()
    return {"id": cur.lastrowid}

@app.get("/api-requests/my")
def my_requests(x_token: str = Header(default="")):
    user = get_current_user(x_token)
    with get_db() as conn:
        rows = conn.execute("""
            SELECT r.*,
                   a.provider, a.base_url,
                   u.username AS manager_username,
                   rv.username AS reviewer_username,
                   ak.name AS api_key_name,
                   (SELECT km.username FROM api_keys k
                      JOIN sub_accounts s ON s.id = k.sub_account_id
                      JOIN users km ON km.id = k.manager_user_id
                      WHERE s.account_id = r.account_id AND k.is_active = 1
                      ORDER BY k.id LIMIT 1) AS key_manager_username,
                   c.name as config_name, c.provider as config_provider, c.manager as config_manager
            FROM api_requests r
            LEFT JOIN accounts a ON a.id = r.account_id
            LEFT JOIN users u ON u.id = a.manager_user_id
            LEFT JOIN users rv ON rv.id = r.reviewer_id
            LEFT JOIN api_configs c ON c.id = r.config_id
            LEFT JOIN api_keys ak ON ak.id = r.api_key_id
            WHERE r.user_id=?
            ORDER BY r.created_at DESC
        """, (user["id"],)).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        d["provider"] = d["provider"] or d.get("config_provider") or ""
        d["manager"] = d["reviewer_username"] or d.get("key_manager_username") or ""

        d["config_name"] = d["config_name"] or ""
        result.append(d)
    return result


# ── Admin: 申请审核 ────────────────────────────────────────

class ReviewIn(BaseModel):
    status: Literal["approved", "rejected"]
    review_note: str = ""
    api_key_id: int | None = None

@app.get("/admin/api-requests")
def admin_list_requests(x_token: str = Header(default="")):
    user = get_current_user(x_token)
    with get_db() as conn:
        if user["role"] == "admin":
            rows = conn.execute("""
                SELECT r.*, u.username,
                       a.provider, a.base_url,
                       mgr.username AS manager_username,
                       rv.username AS reviewer_username,
                       ak.name AS api_key_name,
                       (SELECT km.username FROM api_keys k
                          JOIN sub_accounts s ON s.id = k.sub_account_id
                          JOIN users km ON km.id = k.manager_user_id
                          WHERE s.account_id = r.account_id AND k.is_active = 1
                          ORDER BY k.id LIMIT 1) AS key_manager_username,
                       c.name as config_name, c.provider as config_provider, c.manager as config_manager
                FROM api_requests r
                JOIN users u ON u.id = r.user_id
                LEFT JOIN accounts a ON a.id = r.account_id
                LEFT JOIN users mgr ON mgr.id = a.manager_user_id
                LEFT JOIN users rv ON rv.id = r.reviewer_id
                LEFT JOIN api_configs c ON c.id = r.config_id
                LEFT JOIN api_keys ak ON ak.id = r.api_key_id
                ORDER BY r.created_at DESC
            """).fetchall()
        else:
            # 非 admin：只看分配给自己审核的申请（reviewer_id = self），
            # 或自己是该 account 下任意 key 的管理员（兼容旧数据 reviewer_id 为空）
            rows = conn.execute("""
                SELECT r.*, u.username,
                       a.provider, a.base_url,
                       mgr.username AS manager_username,
                       rv.username AS reviewer_username,
                       ak.name AS api_key_name,
                       (SELECT km.username FROM api_keys k
                          JOIN sub_accounts s ON s.id = k.sub_account_id
                          JOIN users km ON km.id = k.manager_user_id
                          WHERE s.account_id = r.account_id AND k.is_active = 1
                          ORDER BY k.id LIMIT 1) AS key_manager_username,
                       c.name as config_name, c.provider as config_provider, c.manager as config_manager
                FROM api_requests r
                JOIN users u ON u.id = r.user_id
                LEFT JOIN accounts a ON a.id = r.account_id
                LEFT JOIN users mgr ON mgr.id = a.manager_user_id
                LEFT JOIN users rv ON rv.id = r.reviewer_id
                LEFT JOIN api_configs c ON c.id = r.config_id
                LEFT JOIN api_keys ak ON ak.id = r.api_key_id
                WHERE r.reviewer_id = ?
                ORDER BY r.created_at DESC
            """, (user["id"],)).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        d["provider"] = d["provider"] or d.get("config_provider") or ""
        d["manager"] = d.get("reviewer_username") or d.get("key_manager_username") or ""

        d["config_name"] = d["config_name"] or ""
        result.append(d)
    return result


# ── Admin: 用户管理 ───────────────────────────────────────

@app.get("/admin/users")
def list_users(x_token: str = Header(default="")):
    get_current_user(x_token)  # 登录即可
    with get_db() as conn:
        rows = conn.execute(
            "SELECT id, username, role, created_at FROM users ORDER BY id"
        ).fetchall()
    return [dict(r) for r in rows]

class UserRoleIn(BaseModel):
    role: Literal["user", "admin"]

@app.put("/admin/users/{user_id}/role")
def update_user_role(user_id: int, body: UserRoleIn, x_admin_password: str = Header(default="")):
    require_admin(x_admin_password)
    with get_db() as conn:
        row = conn.execute("SELECT id FROM users WHERE id=?", (user_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="用户不存在")
        conn.execute("UPDATE users SET role=? WHERE id=?", (body.role, user_id))
        conn.commit()
    return {"ok": True}

class ResetPwdIn(BaseModel):
    password: str

@app.put("/admin/users/{user_id}/password")
def reset_user_password(user_id: int, body: ResetPwdIn, x_admin_password: str = Header(default="")):
    require_admin(x_admin_password)
    if len(body.password) < 6:
        raise HTTPException(status_code=400, detail="密码至少6位")
    with get_db() as conn:
        row = conn.execute("SELECT id FROM users WHERE id=?", (user_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="用户不存在")
        conn.execute("UPDATE users SET password=? WHERE id=?", (hash_password(body.password), user_id))
        conn.commit()
    return {"ok": True}

@app.delete("/admin/users/{user_id}")
def delete_user(user_id: int, x_admin_password: str = Header(default="")):
    require_admin(x_admin_password)
    with get_db() as conn:
        # 不允许删除最后一个管理员
        admin_count = conn.execute("SELECT COUNT(*) FROM users WHERE role='admin'").fetchone()[0]
        user = conn.execute("SELECT role FROM users WHERE id=?", (user_id,)).fetchone()
        if user and user["role"] == "admin" and admin_count <= 1:
            raise HTTPException(status_code=400, detail="不能删除唯一的管理员账号")
        conn.execute("DELETE FROM user_tokens WHERE user_id=?", (user_id,))
        conn.execute("DELETE FROM users WHERE id=?", (user_id,))
        conn.commit()
    return {"ok": True}

# ── 对话历史 ──────────────────────────────────────────────

class SessionIn(BaseModel):
    config_id: int | None = None
    model: str = ""
    title: str = ""

class MessageIn(BaseModel):
    session_id: int
    role: Literal["user", "assistant"]
    content: str
    model: str = ""

@app.post("/sessions")
def create_session(body: SessionIn, x_token: str = Header(default="")):
    user = get_current_user(x_token)
    now = datetime.now(timezone.utc).isoformat()
    with get_db() as conn:
        cur = conn.execute(
            "INSERT INTO chat_sessions (user_id, config_id, model, title, created_at) VALUES (?,?,?,?,?)",
            (user["id"], body.config_id, body.model, body.title, now)
        )
        conn.commit()
    return {"session_id": cur.lastrowid}

class SessionRenameIn(BaseModel):
    title: str

@app.put("/sessions/{session_id}/title")
def rename_session(session_id: int, body: SessionRenameIn, x_token: str = Header(default="")):
    user = get_current_user(x_token)
    with get_db() as conn:
        row = conn.execute("SELECT user_id FROM chat_sessions WHERE id=?", (session_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="会话不存在")
        if row["user_id"] != user["id"] and user["role"] != "admin":
            raise HTTPException(status_code=403, detail="无权限")
        conn.execute("UPDATE chat_sessions SET title=? WHERE id=?", (body.title.strip(), session_id))
        conn.commit()
    return {"ok": True}

@app.put("/sessions/{session_id}/pin")
def pin_session(session_id: int, x_token: str = Header(default="")):
    user = get_current_user(x_token)
    with get_db() as conn:
        row = conn.execute("SELECT user_id, pinned FROM chat_sessions WHERE id=?", (session_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="会话不存在")
        if row["user_id"] != user["id"] and user["role"] != "admin":
            raise HTTPException(status_code=403, detail="无权限")
        new_pinned = 0 if row["pinned"] else 1
        conn.execute("UPDATE chat_sessions SET pinned=? WHERE id=?", (new_pinned, session_id))
        conn.commit()
    return {"pinned": bool(new_pinned)}

@app.delete("/sessions/{session_id}")
def delete_session(session_id: int, x_token: str = Header(default="")):
    user = get_current_user(x_token)
    with get_db() as conn:
        row = conn.execute("SELECT user_id FROM chat_sessions WHERE id=?", (session_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="会话不存在")
        if row["user_id"] != user["id"] and user["role"] != "admin":
            raise HTTPException(status_code=403, detail="无权限")
        conn.execute("DELETE FROM chat_messages WHERE session_id=?", (session_id,))
        conn.execute("DELETE FROM chat_sessions WHERE id=?", (session_id,))
        conn.commit()
    return {"ok": True}

@app.post("/sessions/messages")
def save_message(body: MessageIn, x_token: str = Header(default="")):
    get_current_user(x_token)
    now = datetime.now(timezone.utc).isoformat()
    # content 若为列表，序列化为 JSON 字符串完整保存（含图片 base64）
    import json as _json
    content = body.content if isinstance(body.content, str) else _json.dumps(body.content, ensure_ascii=False)
    with get_db() as conn:
        conn.execute(
            "INSERT INTO chat_messages (session_id, role, content, model, created_at) VALUES (?,?,?,?,?)",
            (body.session_id, body.role, content, body.model, now)
        )
        conn.commit()
    return {"ok": True}

@app.get("/history/my")
def my_history(x_token: str = Header(default="")):
    user = get_current_user(x_token)
    with get_db() as conn:
        rows = conn.execute("""
            SELECT s.*, c.name as config_name,
                   (SELECT COUNT(*) FROM chat_messages WHERE session_id=s.id) as msg_count
            FROM chat_sessions s
            LEFT JOIN api_configs c ON c.id = s.config_id
            WHERE s.user_id = ?
            ORDER BY s.pinned DESC, s.created_at DESC LIMIT 100
        """, (user["id"],)).fetchall()
    return [dict(r) for r in rows]

@app.get("/history/{session_id}")
def get_session_messages(session_id: int, x_token: str = Header(default="")):
    user = get_current_user(x_token)
    with get_db() as conn:
        session = conn.execute(
            "SELECT * FROM chat_sessions WHERE id=?", (session_id,)
        ).fetchone()
        if not session:
            raise HTTPException(status_code=404, detail="会话不存在")
        # 仅本人或管理员可查看
        if session["user_id"] != user["id"] and user["role"] != "admin":
            raise HTTPException(status_code=403, detail="无权限")
        msgs = conn.execute(
            "SELECT * FROM chat_messages WHERE session_id=? ORDER BY created_at",
            (session_id,)
        ).fetchall()
    return {"session": dict(session), "messages": [dict(m) for m in msgs]}

@app.get("/admin/history")
def admin_history(user_id: int | None = None, x_admin_password: str = Header(default="")):
    require_admin(x_admin_password)
    with get_db() as conn:
        if user_id:
            rows = conn.execute("""
                SELECT s.*, u.username, c.name as config_name,
                       (SELECT COUNT(*) FROM chat_messages WHERE session_id=s.id) as msg_count
                FROM chat_sessions s
                JOIN users u ON u.id = s.user_id
                LEFT JOIN api_configs c ON c.id = s.config_id
                WHERE s.user_id = ?
                ORDER BY s.created_at DESC LIMIT 200
            """, (user_id,)).fetchall()
        else:
            rows = conn.execute("""
                SELECT s.*, u.username, c.name as config_name,
                       (SELECT COUNT(*) FROM chat_messages WHERE session_id=s.id) as msg_count
                FROM chat_sessions s
                JOIN users u ON u.id = s.user_id
                LEFT JOIN api_configs c ON c.id = s.config_id
                ORDER BY s.created_at DESC LIMIT 200
            """).fetchall()
    return [dict(r) for r in rows]

# Spec 1: 旧子账号端点(/admin/configs/{id}/sub-accounts, /admin/sub-accounts/*)
# 已移除,改用 /admin/accounts/{id}/sub-accounts + /admin/sub-accounts/{id}


@app.get("/admin/api-requests/{req_id}/candidate-keys")
def candidate_keys(req_id: int, x_token: str = Header(default="")):
    """列出该申请对应 account 下所有 sub_account 的未占用 key（含余额）。"""
    user = get_current_user(x_token)
    with get_db() as conn:
        req = conn.execute("SELECT * FROM api_requests WHERE id=?", (req_id,)).fetchone()
        if not req:
            raise HTTPException(404, "申请不存在")
        if user["role"] != "admin" and req["reviewer_id"] != user["id"]:
            is_key_manager = conn.execute("""
                SELECT 1 FROM api_keys k
                JOIN sub_accounts s ON s.id = k.sub_account_id
                WHERE s.account_id = ? AND k.manager_user_id = ? AND k.is_active = 1
                LIMIT 1
            """, (req["account_id"], user["id"])).fetchone()
            if not is_key_manager:
                raise HTTPException(403, "无权操作")
        account_id = req["account_id"]
        if not account_id:
            # 回退：用 reviewer_id 反查其管理的 api_keys（key 级管理员），
            # 兼容老申请（无 account_id）和新流程（管理员归 key 不归 account）
            reviewer_id = req["reviewer_id"]
            if not reviewer_id:
                return []
            rows = conn.execute("""
                SELECT sa.id AS sub_account_id, sa.name AS sub_account_name,
                       k.id AS api_key_id, k.name AS api_key_name,
                       k.last_total, k.last_balance
                FROM api_keys k
                JOIN sub_accounts sa ON sa.id = k.sub_account_id
                WHERE k.manager_user_id = ?
                  AND k.is_active = 1
                  AND NOT EXISTS (
                    SELECT 1 FROM api_requests r2
                    WHERE r2.api_key_id = k.id AND r2.status = 'approved'
                  )
                ORDER BY sa.name, k.name
            """, (reviewer_id,)).fetchall()
            return [dict(r) for r in rows]
        rows = conn.execute("""
            SELECT sa.id AS sub_account_id, sa.name AS sub_account_name,
                   k.id AS api_key_id, k.name AS api_key_name,
                   k.last_total, k.last_balance
            FROM sub_accounts sa
            JOIN api_keys k ON k.sub_account_id = sa.id
            WHERE sa.account_id = ?
              AND k.is_active = 1
              AND NOT EXISTS (
                SELECT 1 FROM api_requests r2
                WHERE r2.api_key_id = k.id AND r2.status = 'approved'
              )
            ORDER BY sa.name, k.name
        """, (account_id,)).fetchall()
    return [dict(r) for r in rows]

@app.put("/admin/api-requests/{req_id}")
def admin_review_request(req_id: int, body: ReviewIn, x_token: str = Header(default="")):
    """审核申请；通过时必须传 api_key_id，执行 1:1 占用约束。"""
    user = get_current_user(x_token)
    now = datetime.now(timezone.utc).isoformat()
    with get_db() as conn:
        req = conn.execute("SELECT * FROM api_requests WHERE id=?", (req_id,)).fetchone()
        if not req:
            raise HTTPException(404, "申请不存在")
        if user["role"] != "admin" and req["reviewer_id"] != user["id"]:
            is_key_manager = conn.execute("""
                SELECT 1 FROM api_keys k
                JOIN sub_accounts s ON s.id = k.sub_account_id
                WHERE s.account_id = ? AND k.manager_user_id = ? AND k.is_active = 1
                LIMIT 1
            """, (req["account_id"], user["id"])).fetchone()
            if not is_key_manager:
                raise HTTPException(403, "无权操作")
        if body.status == "approved" and body.api_key_id:
            # 1:1 约束：同一 key 不能同时 approved 给两个申请
            conflict = conn.execute(
                "SELECT COUNT(*) FROM api_requests WHERE api_key_id=? AND status='approved' AND id!=?",
                (body.api_key_id, req_id)
            ).fetchone()[0]
            if conflict > 0:
                raise HTTPException(409, "该密钥已被其他申请占用")
            conn.execute(
                "UPDATE api_requests SET status=?,review_note=?,api_key_id=?,updated_at=? WHERE id=?",
                (body.status, body.review_note, body.api_key_id, now, req_id)
            )
        else:
            conn.execute(
                "UPDATE api_requests SET status=?,review_note=?,updated_at=? WHERE id=?",
                (body.status, body.review_note, now, req_id)
            )
        conn.commit()
    return {"ok": True}

def get_credentials_by_request(request_id: int, user_id: int):
    """反查 api_requests → api_keys → accounts，校验归属和状态，返回 (api_key, base_url, provider, api_key_id)。"""
    with get_db() as conn:
        row = conn.execute("""
            SELECT r.status, r.user_id, k.id AS api_key_id, k.api_key, a.base_url, a.provider
            FROM api_requests r
            JOIN api_keys k ON k.id = r.api_key_id
            JOIN sub_accounts sa ON sa.id = k.sub_account_id
            JOIN accounts a ON a.id = sa.account_id
            WHERE r.id = ?
        """, (request_id,)).fetchone()
    if not row:
        raise HTTPException(404, "申请不存在或未分配密钥")
    if row["user_id"] != user_id:
        raise HTTPException(403, "无权使用此申请")
    if row["status"] != "approved":
        raise HTTPException(403, "申请未通过审核")
    return row["api_key"], row["base_url"], row["provider"], row["api_key_id"]

@app.get("/api-requests/approved")
def approved_requests(x_token: str = Header(default="")):
    """返回当前用户已审核通过的申请列表，不暴露 api_key 明文。"""
    user = get_current_user(x_token)
    with get_db() as conn:
        rows = conn.execute("""
            SELECT r.id AS request_id, r.project_name, r.created_at, r.status,
                   a.provider, a.base_url,
                   sa.name AS sub_account_name,
                   k.name AS api_name,
                   a.models AS available_models
            FROM api_requests r
            LEFT JOIN accounts a ON a.id = r.account_id
            LEFT JOIN api_keys k ON k.id = r.api_key_id
            LEFT JOIN sub_accounts sa ON sa.id = k.sub_account_id
            WHERE r.user_id=? AND r.status='approved'
            ORDER BY r.created_at DESC
        """, (user["id"],)).fetchall()
    return [dict(r) for r in rows]

# ── SQLite 批量历史 ────────────────────────────────────────

class BatchJobIn(BaseModel):
    task_name: str
    model: str = ""
    config_id: int | None = None
    label: str = ""
    row_count: int = 0
    settings_json: str = ""

class BatchJobRowIn(BaseModel):
    job_id: int
    row_index: int
    input_json: str
    output_text: str = ""
    output_type: str = "text"
    output_path: str = ""
    label: str = ""
    success: int = 1
    error_msg: str = ""
    started_at: str = ""
    finished_at: str = ""

@app.post("/batch2/jobs")
def create_batch_job(body: BatchJobIn, x_token: str = Header(default="")):
    user = get_current_user(x_token)
    now_local = datetime.now()
    now_iso   = datetime.now(timezone.utc).isoformat()
    stamp     = now_local.strftime("%Y%m%d_%H%M%S")
    # 同秒内递增序号
    with get_db() as conn:
        existing = conn.execute(
            "SELECT batch_id FROM batch_jobs WHERE batch_id LIKE ?",
            (f"B_{stamp}_%",)
        ).fetchall()
        seq = 1
        for r in existing:
            try:
                n = int((r["batch_id"] or "").rsplit("_", 1)[-1])
                if n >= seq:
                    seq = n + 1
            except Exception:
                pass
        batch_id = f"B_{stamp}_{seq}"
        task_name = body.task_name.strip() if body.task_name and body.task_name.strip() else "NA"
        cur = conn.execute(
            "INSERT INTO batch_jobs (user_id,task_name,model,config_id,label,row_count,settings_json,created_at,started_at,batch_id,status) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            (user["id"], task_name, body.model, body.config_id, body.label, body.row_count,
             body.settings_json, now_iso, now_iso, batch_id, "running")
        )
        conn.commit()
    return {"job_id": cur.lastrowid, "batch_id": batch_id, "task_name": task_name}

@app.post("/batch2/rows")
def save_batch_row(body: BatchJobRowIn, x_token: str = Header(default="")):
    get_current_user(x_token)
    with get_db() as conn:
        conn.execute(
            "INSERT INTO batch_job_rows (job_id,row_index,input_json,output_text,output_type,output_path,label,success,error_msg,started_at,finished_at) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            (body.job_id, body.row_index, body.input_json, body.output_text, body.output_type, body.output_path,
             body.label, body.success, body.error_msg, body.started_at, body.finished_at)
        )
        if body.success:
            conn.execute("UPDATE batch_jobs SET done_count=done_count+1 WHERE id=?", (body.job_id,))
        else:
            conn.execute("UPDATE batch_jobs SET fail_count=fail_count+1 WHERE id=?", (body.job_id,))
        # 让 row_count 至少等于已写入的最大 row_index+1，避免详情页"总数=0"
        conn.execute(
            "UPDATE batch_jobs SET row_count=MAX(row_count, ?) WHERE id=?",
            (int(body.row_index) + 1, body.job_id)
        )
        conn.commit()
    return {"ok": True}

@app.put("/batch2/jobs/{job_id}/finish")
def finish_batch_job(job_id: int, status: str = "", x_token: str = Header(default="")):
    """status 不传时由 done/fail 计数自动判定。"""
    user = get_current_user(x_token)
    now_iso = datetime.now(timezone.utc).isoformat()
    with get_db() as conn:
        row = conn.execute("SELECT user_id,row_count,done_count,fail_count FROM batch_jobs WHERE id=?",
                            (job_id,)).fetchone()
        if not row or row["user_id"] != user["id"]:
            raise HTTPException(status_code=403, detail="无权限")
        if not status:
            done = row["done_count"] or 0
            fail = row["fail_count"] or 0
            total = row["row_count"] or (done + fail)
            if fail == 0 and done >= total:
                status = "completed"
            elif done == 0 and fail > 0:
                status = "failed"
            elif fail > 0:
                status = "partial_failed"
            else:
                status = "completed"
        conn.execute("UPDATE batch_jobs SET status=?, finished_at=? WHERE id=?",
                     (status, now_iso, job_id))
        conn.commit()
    return {"ok": True, "status": status}

class TaskRenameIn(BaseModel):
    task_name: str

@app.put("/batch2/jobs/{job_id}/rename")
def rename_batch_job(job_id: int, body: TaskRenameIn, x_token: str = Header(default="")):
    user = get_current_user(x_token)
    new_name = (body.task_name or "").strip() or "NA"
    with get_db() as conn:
        row = conn.execute("SELECT user_id FROM batch_jobs WHERE id=?", (job_id,)).fetchone()
        if not row or row["user_id"] != user["id"]:
            raise HTTPException(status_code=403, detail="无权限")
        conn.execute("UPDATE batch_jobs SET task_name=? WHERE id=?", (new_name, job_id))
        conn.commit()
    return {"ok": True, "task_name": new_name}

# ── 本地图片预览（白名单：仅当前用户某条批次行 input_json 出现过的路径）────
_IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp"}

@app.get("/batch2/local-image")
def serve_local_image(path: str, token: str = "", x_token: str = Header(default="")):
    user = get_current_user(x_token or token)
    path = (path or "").strip()
    if not path:
        raise HTTPException(status_code=400, detail="path 不能为空")
    ext = pathlib.Path(path).suffix.lower()
    if ext not in _IMAGE_EXTS:
        raise HTTPException(status_code=400, detail="仅支持图片文件")
    if not os.path.isfile(path):
        raise HTTPException(status_code=404, detail="文件不存在")

    # 白名单：路径必须在当前用户的某个 batch_job_rows.input_json 里出现过
    with get_db() as conn:
        rows = conn.execute("""
            SELECT br.input_json FROM batch_job_rows br
            JOIN batch_jobs bj ON bj.id = br.job_id
            WHERE bj.user_id=? AND br.input_json LIKE ?
            LIMIT 1
        """, (user["id"], f"%{path}%")).fetchone()
    if not rows:
        raise HTTPException(status_code=403, detail="该路径未出现在你的批次记录中")

    mime_map = {".jpg":"image/jpeg",".jpeg":"image/jpeg",".png":"image/png",
                ".gif":"image/gif",".webp":"image/webp",".bmp":"image/bmp"}
    from fastapi.responses import FileResponse
    return FileResponse(path, media_type=mime_map.get(ext, "image/jpeg"))

@app.get("/batch2/jobs/{job_id}/export")
def export_batch_job(job_id: int, token: str = "", x_token: str = Header(default="")):
    user = get_current_user(x_token or token)
    with get_db() as conn:
        job = conn.execute("SELECT * FROM batch_jobs WHERE id=?", (job_id,)).fetchone()
        if not job or job["user_id"] != user["id"]:
            raise HTTPException(status_code=403, detail="无权限")
        rows = conn.execute(
            "SELECT * FROM batch_job_rows WHERE job_id=? ORDER BY row_index", (job_id,)
        ).fetchall()

    settings = {}
    try: settings = json.loads(job["settings_json"] or "{}")
    except Exception: pass
    fields = settings.get("selected_fields") or []

    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active; ws.title = "结果"
    headers = ["行号"] + (fields if fields else ["输入"]) + ["输出", "状态", "耗时(ms)", "错误"]
    ws.append(headers)
    for r in rows:
        try: inp = json.loads(r["input_json"] or "{}")
        except Exception: inp = {}
        if fields:
            in_cells = [str(inp.get(f, "")) for f in fields]
        else:
            in_cells = [json.dumps(inp, ensure_ascii=False)]
        # 计算耗时（ms）
        elapsed_ms = ""
        try:
            if r["started_at"] and r["finished_at"]:
                s = datetime.fromisoformat(r["started_at"])
                e = datetime.fromisoformat(r["finished_at"])
                elapsed_ms = int((e - s).total_seconds() * 1000)
        except Exception: pass
        ws.append([r["row_index"]] + in_cells +
                  [r["output_text"] or "", "完成" if r["success"] else "失败",
                   elapsed_ms, r["error_msg"] or ""])

    import io
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    safe_name = (job["batch_id"] or f"job_{job_id}").replace("/", "_")
    from fastapi.responses import Response
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={safe_name}.xlsx"}
    )

# ── 任务（Task）：可视化配置 + 脚本的容器 ─────────────────────

def _gen_script_from_config(cfg: dict) -> str:
    """根据点击配置 config_json 生成 Python 脚本模板。"""
    source     = (cfg.get("source") or "file").lower()
    input_file = cfg.get("input_file", "") or ""
    dataset_id = cfg.get("dataset_id", "") or ""
    fields     = cfg.get("selected_fields") or []
    template   = cfg.get("prompt_template", "") or ""
    # 老 job 没 source 字段时,只要有 dataset_id 就当 dataset
    if source not in ("file", "dataset"):
        source = "dataset" if dataset_id else "file"
    # 反斜杠用 raw 字符串规避
    fields_py = json.dumps(fields, ensure_ascii=False)
    template_py = json.dumps(template, ensure_ascii=False)
    if source == "dataset":
        load_block = (
            f'DATASET_ID_LOCAL = "{dataset_id}"\n'
            "df = load_dataset(DATASET_ID_LOCAL)\n"
            'print(f"已加载 {len(df)} 行 (数据集 " + DATASET_ID_LOCAL + ")")'
        )
    else:
        load_block = (
            f'INPUT_FILE = r"{input_file}"\n'
            'df = pd.read_excel(INPUT_FILE) if INPUT_FILE.endswith((".xlsx", ".xls")) else (\n'
            '    pd.read_csv(INPUT_FILE) if INPUT_FILE.endswith(".csv") else\n'
            '    pd.read_json(INPUT_FILE) if INPUT_FILE.endswith(".json") else\n'
            '    pd.DataFrame()\n'
            ')\n'
            'print(f"已加载 {len(df)} 行")'
        )
    return f'''# 自动生成的批跑脚本（来自点击配置）
# 在「点击配置」修改任何字段会实时重生成此代码,除非你手动编辑过
import os, json, pandas as pd

SELECTED_FIELDS  = {fields_py}
PROMPT_TEMPLATE  = {template_py}

{load_block}

for i, row in df.iterrows():
    if i in _RESUME_DONE:    # 续跑时跳过已完成行
        continue
    inp = {{f: row.get(f, "") for f in SELECTED_FIELDS}}
    prompt = PROMPT_TEMPLATE
    image_paths = []
    # 对所有列做 {{field}} 替换,不限于 SELECTED_FIELDS,
    # 避免用户在 prompt 里写了未点击 chip 的字段时未被替换
    for col in row.index:
        v = row[col]
        sv = "" if (v is None or (isinstance(v, float) and pd.isna(v))) else str(v)
        if isinstance(sv, str) and sv.lower().endswith((".png",".jpg",".jpeg",".gif",".webp")) and os.path.exists(sv):
            if sv not in image_paths:
                image_paths.append(sv)
            prompt = prompt.replace("{{" + str(col) + "}}", "（见附图）")
        else:
            prompt = prompt.replace("{{" + str(col) + "}}", sv)
    try:
        if image_paths:
            content = [{{"image": p}} for p in image_paths] + [prompt]
            result = chat(content)
        else:
            result = chat(prompt)
        record_row(i, input=inp, output=result, success=True)
        df.at[i, "AI结果"] = result
    except Exception as e:
        record_row(i, input=inp, output="", success=False, error=str(e))

output_path = os.path.join(WORK_DIR, "结果.xlsx")
df.to_excel(output_path, index=False)
save_file(output_path)
'''


class LintIn(BaseModel):
    code: str

@app.post("/batch2/lint")
def lint_script(body: LintIn, x_token: str = Header(default="")):
    get_current_user(x_token)
    import ast as _ast
    try:
        _ast.parse(body.code or "")
        return {"ok": True}
    except SyntaxError as e:
        return {"ok": False, "error": f"L{e.lineno} C{e.offset}: {e.msg}"}


class BatchRunIn(BaseModel):
    source_type: str = "click"          # 'click' | 'script'
    task_name: str = ""
    config_json: str = "{}"             # 点击配置：保存表单；脚本：可空
    script_code: str = ""               # 脚本：用户脚本；点击配置：可空（由 config 自动生成）
    request_id: int | None = None
    model: str = ""
    resume_job_id: int | None = None    # 续跑同一 job
    rerun_from_job_id: int | None = None  # 基于历史 job 重跑（取其字段）

def _next_batch_id(conn) -> str:
    now_local = datetime.now()
    stamp = now_local.strftime("%Y%m%d_%H%M%S")
    existing = conn.execute(
        "SELECT batch_id FROM batch_jobs WHERE batch_id LIKE ?", (f"B_{stamp}_%",)
    ).fetchall()
    seq = 1
    for r in existing:
        try:
            n = int((r["batch_id"] or "").rsplit("_", 1)[-1])
            if n >= seq: seq = n + 1
        except Exception: pass
    return f"B_{stamp}_{seq}"

@app.post("/batch2/run")
def batch_run(body: BatchRunIn, x_token: str = Header(default="")):
    """统一启动入口：点击配置 / 脚本两种来源都走这里。
    返回 SSE 流。每条 batch_jobs 自带完整 config_json + script_code,可独立重跑。"""
    user = get_current_user(x_token)

    source_type = (body.source_type or "click").lower()
    if source_type not in ("click", "script"):
        raise HTTPException(status_code=400, detail="source_type 必须为 click 或 script")

    cfg = {}
    try: cfg = json.loads(body.config_json or "{}")
    except Exception: cfg = {}

    task_name  = (body.task_name or "").strip()  # 空时下文用 batch_id 兜底
    request_id = body.request_id or cfg.get("request_id")
    model      = body.model or cfg.get("model") or ""

    # 决定运行的代码
    if source_type == "click":
        code = _gen_script_from_config(cfg)
    else:
        code = (body.script_code or "").strip()
        if not code:
            raise HTTPException(status_code=400, detail="脚本不能为空")

    now_iso = datetime.now(timezone.utc).isoformat()
    resume_done_csv = ""
    job_id = None
    batch_id = None

    if body.resume_job_id:
        with get_db() as conn:
            job_row = conn.execute("SELECT * FROM batch_jobs WHERE id=?",
                                   (body.resume_job_id,)).fetchone()
            if not job_row or job_row["user_id"] != user["id"]:
                raise HTTPException(status_code=403, detail="无权限")
            done_idx = conn.execute(
                "SELECT row_index FROM batch_job_rows WHERE job_id=? AND success=1",
                (body.resume_job_id,)
            ).fetchall()
            resume_done_csv = ",".join(str(r["row_index"]) for r in done_idx)
            conn.execute("UPDATE batch_jobs SET status='running' WHERE id=?",
                         (body.resume_job_id,))
            conn.commit()
        job_id   = body.resume_job_id
        batch_id = job_row["batch_id"]
        # 续跑：用 job 自身存的代码/配置
        try:
            code = job_row["script_code"] or code
            cfg  = json.loads(job_row["config_json"] or "{}") or cfg
        except Exception: pass
        source_type = job_row["source_type"] or source_type
        task_name   = job_row["task_name"] or task_name
        request_id  = request_id or cfg.get("request_id")
        model       = model or job_row["model"]
    else:
        with get_db() as conn:
            batch_id = _next_batch_id(conn)
            # 任务名留空时,直接复用 batch_id —— 天然唯一,符合"同任务ID命名规则"
            if not task_name:
                task_name = batch_id
            dup = conn.execute(
                "SELECT id FROM batch_jobs WHERE user_id=? AND task_name=? LIMIT 1",
                (user["id"], task_name)
            ).fetchone()
            if dup:
                raise HTTPException(
                    status_code=409,
                    detail=f"任务名「{task_name}」已存在，请换一个名字"
                )
            settings = {
                "source_type": source_type,
                "task_name":   task_name,
                "model":       model,
                "request_id":  request_id,
                "config":      cfg,
            }
            cur = conn.execute(
                "INSERT INTO batch_jobs (user_id,task_name,model,config_id,label,row_count,"
                "settings_json,created_at,started_at,batch_id,status,source_type,"
                "script_code,config_json) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (user["id"], task_name, model, None, "", 0,
                 json.dumps(settings, ensure_ascii=False), now_iso, now_iso,
                 batch_id, "running", source_type, code, json.dumps(cfg, ensure_ascii=False))
            )
            job_id = cur.lastrowid
            conn.commit()

    run_id   = str(uuid.uuid4())
    work_dir = os.path.join(SCRIPT_BASE_DIR, run_id)
    os.makedirs(work_dir, exist_ok=True)
    script_path = os.path.join(work_dir, "user_script.py")
    with open(script_path, "w", encoding="utf-8") as f:
        f.write(code)

    cmd = [
        sys.executable, SCRIPT_WORKER,
        "--run-id",     run_id,
        "--script",     script_path,
        "--work-dir",   work_dir,
        "--token",      x_token,
        "--config-id",  "",
        "--request-id", str(request_id) if request_id else "",
        "--model",      model or "",
        "--file-path",  cfg.get("input_file", "") or "",
        "--dataset-id", cfg.get("dataset_id", "") or "",
        "--timeout",    str(SCRIPT_TIMEOUT),
        "--job-id",     str(job_id),
        "--resume-done-csv", resume_done_csv,
    ]

    async def generate():
        proc = await asyncio.create_subprocess_exec(
            *cmd, stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE,
        )
        _RUNNING_PROCS[job_id] = proc
        yield f"data: {json.dumps({'type':'task_started','job_id':job_id,'batch_id':batch_id,'run_id':run_id})}\n\n"
        stdout_buffer = []   # 累积纯文本 stdout（用于脚本类型自动落 row）
        stderr_buffer = []
        try:
            while True:
                try:
                    line = await asyncio.wait_for(
                        proc.stdout.readline(), timeout=SCRIPT_TIMEOUT + 10)
                except asyncio.TimeoutError:
                    proc.kill()
                    yield f"data: {json.dumps({'type':'error','text':'Worker 无响应'})}\n\n"
                    break
                if not line: break
                text = line.decode("utf-8", errors="replace").strip()
                if text:
                    # 尝试解析 worker 的 JSON 行，捕获 stdout/stderr 文本
                    try:
                        evt = json.loads(text)
                        if evt.get("type") == "stdout" and evt.get("text"):
                            stdout_buffer.append(evt["text"])
                        elif evt.get("type") == "stderr" and evt.get("text"):
                            stderr_buffer.append(evt["text"])
                    except Exception:
                        pass
                    yield f"data: {text}\n\n"
            try:
                err = await asyncio.wait_for(proc.stderr.read(), timeout=3)
                if err:
                    msg = err.decode("utf-8", errors="replace").strip()
                    if msg:
                        stderr_buffer.append(msg)
                        yield f"data: {json.dumps({'type':'error','text':msg})}\n\n"
            except asyncio.TimeoutError: pass
        finally:
            try: await proc.wait()
            except Exception: pass
            _RUNNING_PROCS.pop(job_id, None)
            try:
                with get_db() as conn:
                    rr = conn.execute("SELECT row_count,done_count,fail_count,status,source_type FROM batch_jobs WHERE id=?",
                                       (job_id,)).fetchone()
                    if rr and rr["status"] != "paused":
                        done = rr["done_count"] or 0
                        fail = rr["fail_count"] or 0
                        # 脚本类型未调 record_row 时，自动落一条 row 保留日志
                        if (rr["source_type"] == "script") and done == 0 and fail == 0:
                            rc = proc.returncode if proc.returncode is not None else 1
                            success = 1 if rc == 0 else 0
                            output = "".join(stdout_buffer).strip() or "(无 stdout 输出)"
                            err_msg = "".join(stderr_buffer).strip() if not success else ""
                            now_iso = datetime.now(timezone.utc).isoformat()
                            conn.execute(
                                "INSERT INTO batch_job_rows (job_id,row_index,input_json,output_text,output_type,output_path,label,success,error_msg,started_at,finished_at) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                                (job_id, 0,
                                 json.dumps({"脚本": "（用户脚本未调用 record_row）"}, ensure_ascii=False),
                                 output, "text", "", "", success, err_msg, now_iso, now_iso)
                            )
                            if success: conn.execute("UPDATE batch_jobs SET done_count=1 WHERE id=?", (job_id,))
                            else:       conn.execute("UPDATE batch_jobs SET fail_count=1 WHERE id=?", (job_id,))
                            conn.execute("UPDATE batch_jobs SET row_count=MAX(row_count,1) WHERE id=?", (job_id,))
                            done = success; fail = 1 - success

                        if fail == 0 and done > 0: st = "completed"
                        elif done == 0 and fail > 0: st = "failed"
                        elif fail > 0: st = "partial_failed"
                        else: st = "completed"
                        conn.execute("UPDATE batch_jobs SET status=?, finished_at=? WHERE id=?",
                                     (st, datetime.now(timezone.utc).isoformat(), job_id))
                        conn.commit()
            except Exception: pass

    return StreamingResponse(generate(), media_type="text/event-stream")


@app.post("/batch2/jobs/{job_id}/pause")
def pause_job(job_id: int, x_token: str = Header(default="")):
    user = get_current_user(x_token)
    with get_db() as conn:
        row = conn.execute("SELECT user_id FROM batch_jobs WHERE id=?", (job_id,)).fetchone()
        if not row or row["user_id"] != user["id"]:
            raise HTTPException(status_code=403, detail="无权限")
        conn.execute("UPDATE batch_jobs SET status='paused' WHERE id=?", (job_id,))
        conn.commit()
    proc = _RUNNING_PROCS.pop(job_id, None)
    if proc:
        try: proc.kill()
        except Exception: pass
    return {"ok": True}

class JobActionIn(BaseModel):
    request_id: int | None = None
    model: str = ""

@app.post("/batch2/jobs/{job_id}/resume")
def resume_job(job_id: int, body: JobActionIn, x_token: str = Header(default="")):
    """续跑同一个 job，跳过已完成行。"""
    user = get_current_user(x_token)
    with get_db() as conn:
        row = conn.execute("SELECT * FROM batch_jobs WHERE id=?", (job_id,)).fetchone()
        if not row or row["user_id"] != user["id"]:
            raise HTTPException(status_code=403, detail="无权限")
    return batch_run(BatchRunIn(
        source_type   = row["source_type"] or "click",
        task_name     = row["task_name"] or "",
        config_json   = row["config_json"] or "{}",
        script_code   = row["script_code"] or "",
        request_id    = body.request_id,
        model         = body.model or row["model"] or "",
        resume_job_id = job_id,
    ), x_token=x_token)

@app.post("/batch2/jobs/{job_id}/rerun")
def rerun_job(job_id: int, body: JobActionIn, x_token: str = Header(default="")):
    """基于历史 job 的 config + script，开一个全新批次。"""
    user = get_current_user(x_token)
    with get_db() as conn:
        row = conn.execute("SELECT * FROM batch_jobs WHERE id=?", (job_id,)).fetchone()
        if not row or row["user_id"] != user["id"]:
            raise HTTPException(status_code=403, detail="无权限")
    base_name = row["task_name"] or "NA"
    suffix = datetime.now().strftime("%H%M%S")
    new_name = f"{base_name}_rerun{suffix}"
    return batch_run(BatchRunIn(
        source_type = row["source_type"] or "click",
        task_name   = new_name,
        config_json = row["config_json"] or "{}",
        script_code = row["script_code"] or "",
        request_id  = body.request_id,
        model       = body.model or row["model"] or "",
    ), x_token=x_token)


@app.get("/batch2/jobs")
def list_batch_jobs(x_token: str = Header(default="")):
    user = get_current_user(x_token)
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM batch_jobs WHERE user_id=? ORDER BY created_at DESC LIMIT 100",
            (user["id"],)
        ).fetchall()
    return [dict(r) for r in rows]

@app.get("/batch2/jobs/{job_id}/rows")
def get_batch_job_rows(job_id: int, offset: int = 0, limit: int = 0,
                       x_token: str = Header(default="")):
    user = get_current_user(x_token)
    with get_db() as conn:
        job = conn.execute("SELECT user_id FROM batch_jobs WHERE id=?", (job_id,)).fetchone()
        if not job or job["user_id"] != user["id"]:
            raise HTTPException(status_code=403, detail="无权限")
        total = conn.execute("SELECT COUNT(*) AS c FROM batch_job_rows WHERE job_id=?",
                              (job_id,)).fetchone()["c"]
        if limit > 0:
            rows = conn.execute(
                "SELECT * FROM batch_job_rows WHERE job_id=? ORDER BY row_index LIMIT ? OFFSET ?",
                (job_id, limit, offset)
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM batch_job_rows WHERE job_id=? ORDER BY row_index", (job_id,)
            ).fetchall()
    return {"total": total, "rows": [dict(r) for r in rows]} if limit > 0 else [dict(r) for r in rows]

# ── ClickHouse ────────────────────────────────────────────

def get_ch_client():
    try:
        import clickhouse_connect
        return clickhouse_connect.get_client(
            host=os.environ.get("CLICKHOUSE_HOST", "localhost"),
            port=int(os.environ.get("CLICKHOUSE_PORT", "8123")),
            database=os.environ.get("CLICKHOUSE_DB", "default"),
            username=os.environ.get("CLICKHOUSE_USER", "default"),
            password=os.environ.get("CLICKHOUSE_PASSWORD", ""),
        )
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"ClickHouse 连接失败：{e}")

def ensure_ch_tables():
    try:
        ch = get_ch_client()
        ch.command("""
            CREATE TABLE IF NOT EXISTS batch_tasks (
                task_id     String,
                task_name   String,
                label       String,
                row_count   UInt32,
                status      String,
                config_name String,
                settings_json String DEFAULT '',
                created_at  DateTime DEFAULT now()
            ) ENGINE = MergeTree() ORDER BY created_at
        """)
        # 老表兼容：补字段（已存在则忽略）
        try:
            ch.command("ALTER TABLE batch_tasks ADD COLUMN IF NOT EXISTS settings_json String DEFAULT ''")
        except Exception:
            pass
        ch.command("""
            CREATE TABLE IF NOT EXISTS batch_inputs (
                task_id     String,
                row_index   UInt32,
                input_json  String,
                created_at  DateTime DEFAULT now()
            ) ENGINE = MergeTree() ORDER BY (task_id, row_index)
        """)
        ch.command("""
            CREATE TABLE IF NOT EXISTS batch_results (
                task_id     String,
                row_index   UInt32,
                input_json  String,
                output_text String,
                output_type String,
                output_path String,
                label       String,
                model       String,
                success     UInt8,
                error_msg   String,
                created_at  DateTime DEFAULT now()
            ) ENGINE = MergeTree() ORDER BY (task_id, row_index)
        """)
        ch.command("""
            CREATE TABLE IF NOT EXISTS datasets (
                dataset_id  String,
                name        String,
                row_count   UInt32,
                headers     String,
                created_by  String,
                created_at  DateTime DEFAULT now()
            ) ENGINE = MergeTree() ORDER BY created_at
        """)
        ch.command("""
            CREATE TABLE IF NOT EXISTS dataset_rows (
                dataset_id  String,
                row_index   UInt32,
                row_json    String
            ) ENGINE = MergeTree() ORDER BY (dataset_id, row_index)
        """)
    except Exception:
        pass  # ClickHouse 未配置时静默忽略

@app.get("/clickhouse/status")
def ch_status():
    """检查 ClickHouse 连接状态。"""
    try:
        ch = get_ch_client()
        ch.command("SELECT 1")
        return {"connected": True}
    except Exception as e:
        return {"connected": False, "error": str(e)}

# ── 批量处理 ───────────────────────────────────────────────

class BatchStartIn(BaseModel):
    task_name: str
    label: str = ""
    row_count: int
    config_id: int | None = None
    config_name: str = ""
    model: str = ""
    settings_json: str = ""

class BatchRowIn(BaseModel):
    task_id: str
    row_index: int
    input_json: str
    output_text: str = ""
    output_type: str = "text"
    output_path: str = ""
    label: str = ""
    model: str = ""
    success: int = 1
    error_msg: str = ""

@app.post("/batch/start")
def batch_start(body: BatchStartIn, x_token: str = Header(default="")):
    get_current_user(x_token)
    task_id = str(uuid.uuid4())
    try:
        ensure_ch_tables()
        ch = get_ch_client()
        ch.insert("batch_tasks", [[
            task_id, body.task_name, body.label,
            body.row_count, "running", body.config_name, body.settings_json,
        ]], column_names=["task_id","task_name","label","row_count","status","config_name","settings_json"])
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"ClickHouse 写入失败：{e}")
    return {"task_id": task_id}

class BatchInputsIn(BaseModel):
    task_id: str
    rows: list[dict]  # [{row_index: int, input_json: str}, ...]

@app.post("/batch/inputs")
def batch_inputs(body: BatchInputsIn, x_token: str = Header(default="")):
    """批量写入输入数据到 ClickHouse。"""
    get_current_user(x_token)
    try:
        ch = get_ch_client()
        data = [[body.task_id, r["row_index"], r["input_json"]] for r in body.rows]
        ch.insert("batch_inputs", data, column_names=["task_id","row_index","input_json"])
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"ClickHouse 写入失败：{e}")
    return {"ok": True, "count": len(body.rows)}

@app.get("/batch/inputs/{task_id}")
def get_batch_inputs(task_id: str, x_token: str = Header(default="")):
    """从 ClickHouse 读取某任务的输入数据。"""
    get_current_user(x_token)
    try:
        ch = get_ch_client()
        rows = ch.query(
            "SELECT row_index, input_json FROM batch_inputs WHERE task_id=%(tid)s ORDER BY row_index",
            parameters={"tid": task_id}
        )
        return [{"row_index": r[0], "input_json": r[1]} for r in rows.result_rows]
    except Exception as e:
        raise HTTPException(status_code=503, detail=str(e))

@app.post("/batch/row")
def batch_row(body: BatchRowIn, x_token: str = Header(default="")):
    get_current_user(x_token)
    try:
        ch = get_ch_client()
        ch.insert("batch_results", [[
            body.task_id, body.row_index, body.input_json,
            body.output_text, body.output_type, body.output_path,
            body.label, body.model, body.success, body.error_msg
        ]], column_names=[
            "task_id","row_index","input_json","output_text",
            "output_type","output_path","label","model","success","error_msg"
        ])
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"ClickHouse 写入失败：{e}")
    return {"ok": True}

@app.put("/batch/finish/{task_id}")
def batch_finish(task_id: str, x_token: str = Header(default="")):
    get_current_user(x_token)
    try:
        ch = get_ch_client()
        ch.command(f"ALTER TABLE batch_tasks UPDATE status='completed' WHERE task_id='{task_id}'")
    except Exception:
        pass
    return {"ok": True}

@app.get("/batch/tasks")
def batch_tasks(x_token: str = Header(default="")):
    get_current_user(x_token)
    try:
        ch = get_ch_client()
        rows = ch.query("SELECT * FROM batch_tasks ORDER BY created_at DESC LIMIT 100")
        return [dict(zip(rows.column_names, r)) for r in rows.result_rows]
    except Exception as e:
        raise HTTPException(status_code=503, detail=str(e))

@app.get("/batch/results/{task_id}")
def batch_results(task_id: str, x_token: str = Header(default="")):
    get_current_user(x_token)
    try:
        ch = get_ch_client()
        rows = ch.query(
            "SELECT * FROM batch_results WHERE task_id=%(tid)s ORDER BY row_index",
            parameters={"tid": task_id}
        )
        return [dict(zip(rows.column_names, r)) for r in rows.result_rows]
    except Exception as e:
        raise HTTPException(status_code=503, detail=str(e))

# ── 平台数据集（供批跑选用） ───────────────────────────────

class DatasetCreateIn(BaseModel):
    name: str
    rows: list[dict]  # 每行是 {字段: 值, ...}

@app.get("/datasets")
def list_datasets(x_token: str = Header(default="")):
    get_current_user(x_token)
    try:
        ensure_ch_tables()
        ch = get_ch_client()
        rows = ch.query("SELECT dataset_id, name, row_count, headers, created_by, created_at FROM datasets ORDER BY created_at DESC LIMIT 200")
        return [dict(zip(rows.column_names, r)) for r in rows.result_rows]
    except Exception as e:
        raise HTTPException(status_code=503, detail=str(e))

@app.post("/datasets")
def create_dataset(body: DatasetCreateIn, x_token: str = Header(default="")):
    user = get_current_user(x_token)
    if not body.name.strip():
        raise HTTPException(status_code=400, detail="数据集名称不能为空")
    if not body.rows:
        raise HTTPException(status_code=400, detail="数据集为空")
    dataset_id = str(uuid.uuid4())
    headers = list(body.rows[0].keys())
    try:
        ensure_ch_tables()
        ch = get_ch_client()
        ch.insert("datasets", [[
            dataset_id, body.name.strip(), len(body.rows),
            json.dumps(headers, ensure_ascii=False),
            user.get("username", "") if isinstance(user, dict) else "",
        ]], column_names=["dataset_id","name","row_count","headers","created_by"])
        BATCH = 500
        data = [[dataset_id, i, json.dumps(r, ensure_ascii=False)] for i, r in enumerate(body.rows)]
        for b in range(0, len(data), BATCH):
            ch.insert("dataset_rows", data[b:b+BATCH], column_names=["dataset_id","row_index","row_json"])
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"ClickHouse 写入失败：{e}")
    return {"dataset_id": dataset_id, "row_count": len(body.rows)}

@app.get("/datasets/{dataset_id}/rows")
def get_dataset_rows(dataset_id: str, x_token: str = Header(default="")):
    get_current_user(x_token)
    try:
        ch = get_ch_client()
        rows = ch.query(
            "SELECT row_index, row_json FROM dataset_rows WHERE dataset_id=%(did)s ORDER BY row_index",
            parameters={"did": dataset_id}
        )
        out = []
        for ri, rj in rows.result_rows:
            try:
                out.append(json.loads(rj))
            except Exception:
                out.append({})
        return out
    except Exception as e:
        raise HTTPException(status_code=503, detail=str(e))

@app.delete("/datasets/{dataset_id}")
def delete_dataset(dataset_id: str, x_token: str = Header(default="")):
    user = get_current_user(x_token)
    if isinstance(user, dict) and user.get("role") != "admin":
        # 非 admin 只能删自己创建的
        try:
            ch = get_ch_client()
            r = ch.query("SELECT created_by FROM datasets WHERE dataset_id=%(d)s", parameters={"d": dataset_id})
            if r.result_rows and r.result_rows[0][0] != user.get("username"):
                raise HTTPException(status_code=403, detail="无权删除他人创建的数据集")
        except HTTPException:
            raise
        except Exception:
            pass
    try:
        ch = get_ch_client()
        ch.command(f"ALTER TABLE datasets DELETE WHERE dataset_id='{dataset_id}'")
        ch.command(f"ALTER TABLE dataset_rows DELETE WHERE dataset_id='{dataset_id}'")
    except Exception as e:
        raise HTTPException(status_code=503, detail=str(e))
    return {"ok": True}

# ── 保存结果到本地文件 ─────────────────────────────────────

class SaveResultIn(BaseModel):
    path: str
    content: str
    encoding: str = "text"  # "text" | "base64"

@app.post("/save-result")
def save_result(body: SaveResultIn, x_token: str = Header(default="")):
    get_current_user(x_token)
    try:
        # 防御：确保父目录存在
        parent = os.path.dirname(body.path)
        if parent and not os.path.exists(parent):
            os.makedirs(parent, exist_ok=True)
        if body.encoding == "base64":
            import base64
            data = base64.b64decode(body.content)
            with open(body.path, 'wb') as f:
                f.write(data)
        else:
            with open(body.path, 'w', encoding='utf-8') as f:
                f.write(body.content)
        return {"ok": True, "path": body.path}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# ── 脚本编辑器 ────────────────────────────────────────────

import tempfile, threading

SCRIPT_BASE_DIR = os.path.join(tempfile.gettempdir(), "script_runs")
os.makedirs(SCRIPT_BASE_DIR, exist_ok=True)
_RUNNING_PROCS: dict[int, asyncio.subprocess.Process] = {}   # job_id -> proc
SCRIPT_WORKER   = os.path.join(os.path.dirname(os.path.abspath(__file__)), "script_worker.py")
SCRIPT_TIMEOUT  = 600
SCRIPT_MAX_SIZE = 1024 ** 3  # 1 GB

def _check_approved(user_id: int):
    with get_db() as conn:
        row = conn.execute(
            "SELECT id FROM api_requests WHERE user_id=? AND status='approved' LIMIT 1",
            (user_id,)
        ).fetchone()
    if not row:
        raise HTTPException(status_code=403, detail="需要已审批的 API 权限才能运行脚本")

@app.post("/script/upload")
async def script_upload(file: UploadFile, x_token: str = Header(default="")):
    user = get_current_user(x_token)
    _check_approved(user["id"])

    suffix = os.path.splitext(file.filename or "")[1].lower()
    if suffix not in ('.parquet', '.json', '.csv', '.xlsx', '.xls'):
        raise HTTPException(status_code=400, detail="仅支持 parquet / json / csv / xlsx")

    content = await file.read()
    if len(content) > SCRIPT_MAX_SIZE:
        raise HTTPException(status_code=413, detail="文件超过 1GB 上限")

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix, dir=SCRIPT_BASE_DIR)
    tmp.write(content); tmp.close()

    try:
        import pandas as _pd
        loaders = {
            '.parquet': _pd.read_parquet,
            '.json':    _pd.read_json,
            '.xlsx':    _pd.read_excel,
            '.xls':     _pd.read_excel,
            '.csv':     _pd.read_csv,
        }
        df = loaders[suffix](tmp.name)
        rows, cols = df.shape
        col_names  = df.columns.tolist()
    except Exception:
        rows, cols, col_names = 0, 0, []

    return {"file_id": tmp.name, "filename": file.filename,
            "rows": rows, "columns": cols, "col_names": col_names}


@app.post("/script/run")
async def script_run(
    code:       str      = Body(...),
    config_id:  int|None = Body(None),
    request_id: int|None = Body(None),
    model:      str      = Body(""),
    file_id:    str|None = Body(None),
    x_token:    str      = Header(default="")
):
    user = get_current_user(x_token)
    _check_approved(user["id"])

    run_id   = str(uuid.uuid4())
    work_dir = os.path.join(SCRIPT_BASE_DIR, run_id)
    os.makedirs(work_dir, exist_ok=True)

    script_path = os.path.join(work_dir, "user_script.py")
    with open(script_path, "w", encoding="utf-8") as f:
        f.write(code)

    cmd = [
        sys.executable, SCRIPT_WORKER,
        "--run-id",     run_id,
        "--script",     script_path,
        "--work-dir",   work_dir,
        "--token",      x_token,
        "--config-id",  str(config_id) if config_id else "",
        "--request-id", str(request_id) if request_id else "",
        "--model",      model or "",
        "--file-path",  file_id or "",
        "--timeout",    str(SCRIPT_TIMEOUT),
    ]

    async def generate():
        proc = await asyncio.create_subprocess_exec(
            *cmd,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE,
        )
        try:
            while True:
                try:
                    line = await asyncio.wait_for(
                        proc.stdout.readline(), timeout=SCRIPT_TIMEOUT + 10
                    )
                except asyncio.TimeoutError:
                    proc.kill()
                    yield f"data: {json.dumps({'type':'error','text':'Worker 无响应，已强制停止'})}\n\n"
                    break
                if not line:
                    break
                text = line.decode("utf-8", errors="replace").strip()
                if text:
                    yield f"data: {text}\n\n"
            # drain stderr（worker 自身错误）
            try:
                err = await asyncio.wait_for(proc.stderr.read(), timeout=3)
                if err:
                    msg = err.decode("utf-8", errors="replace").strip()
                    if msg:
                        yield f"data: {json.dumps({'type':'error','text':msg})}\n\n"
            except asyncio.TimeoutError:
                pass
        except Exception as e:
            try: proc.kill()
            except Exception: pass
            yield f"data: {json.dumps({'type':'error','text':str(e)})}\n\n"
        finally:
            try: await proc.wait()
            except Exception: pass

    return StreamingResponse(generate(), media_type="text/event-stream")


@app.get("/script/result/{run_id}/download")
def script_download(run_id: str, token: str = "", x_token: str = Header(default="")):
    # 浏览器直接打开 <a href> 无法附带 X-Token header,允许通过 query 传 token
    get_current_user(x_token or token)
    work_dir = os.path.join(SCRIPT_BASE_DIR, run_id)
    if not os.path.isdir(work_dir):
        raise HTTPException(status_code=404, detail="运行结果不存在或已过期")

    import io, zipfile
    from fastapi.responses import Response
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for fname in os.listdir(work_dir):
            fpath = os.path.join(work_dir, fname)
            if os.path.isfile(fpath) and not fname.startswith("_") and fname != "user_script.py":
                zf.write(fpath, fname)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename=result_{run_id[:8]}.zip"}
    )


# ── 后台清理（已禁用：保留所有运行目录,供处理历史下载）────
def _cleanup_old_runs():
    # 用户要求永久保留脚本运行结果,不再自动清理
    return

def _start_cleanup_thread():
    def loop():
        while True:
            time.sleep(3600)
            _cleanup_old_runs()
    t = threading.Thread(target=loop, daemon=True)
    t.start()

_start_cleanup_thread()

# ── 静态文件服务（让前端可通过 http://localhost:8000/ 访问）──
_ROOT = os.path.dirname(os.path.abspath(__file__))
app.mount("/assets", StaticFiles(directory=os.path.join(_ROOT, "assets")), name="assets")
app.mount("/pages",  StaticFiles(directory=os.path.join(_ROOT, "pages")),  name="pages")
app.mount("/config", StaticFiles(directory=os.path.join(_ROOT, "config")), name="config")
if os.path.isdir(os.path.join(_ROOT, "src")):
    app.mount("/src", StaticFiles(directory=os.path.join(_ROOT, "src")), name="src")


# ── 交互式终端（PTY + WebSocket）─────────────────────────────────
def _check_ws_token(token: str) -> dict | None:
    if not token:
        return None
    try:
        with get_db() as conn:
            row = conn.execute(
                "SELECT u.* FROM users u JOIN user_tokens t ON t.user_id=u.id WHERE t.token=?",
                (token,),
            ).fetchone()
        return dict(row) if row else None
    except Exception:
        return None


@app.websocket("/terminal")
async def terminal_ws(ws: WebSocket):
    await ws.accept()
    token = ws.query_params.get("token", "")
    user = _check_ws_token(token)
    if not user:
        await ws.send_text(json.dumps({"type": "error", "msg": "未登录或登录已过期"}))
        await ws.close()
        return

    import pty, fcntl, termios, struct, signal, select, errno

    pid, fd = pty.fork()
    if pid == 0:
        # 子进程：进入项目根目录后启动 shell
        try:
            os.chdir(_ROOT)
        except Exception:
            pass
        env = os.environ.copy()
        env["TERM"] = "xterm-256color"
        env["PS1"] = r"\w $ "
        shell = os.environ.get("SHELL") or "/bin/bash"
        try:
            os.execvpe(shell, [shell, "-l"], env)
        except Exception:
            os.execvpe("/bin/sh", ["/bin/sh"], env)
        return

    def _set_size(rows: int, cols: int):
        try:
            fcntl.ioctl(fd, termios.TIOCSWINSZ,
                        struct.pack("HHHH", rows, cols, 0, 0))
        except Exception:
            pass

    _set_size(30, 100)
    loop = asyncio.get_event_loop()
    closed = False

    async def pty_to_ws():
        nonlocal closed
        while not closed:
            try:
                ready = await loop.run_in_executor(None, select.select, [fd], [], [], 0.5)
                if not ready[0]:
                    if closed:
                        break
                    continue
                data = await loop.run_in_executor(None, os.read, fd, 4096)
                if not data:
                    break
                await ws.send_bytes(data)
            except OSError as e:
                if e.errno == errno.EIO:
                    break
                if closed:
                    break
                continue
            except Exception:
                break
        closed = True

    pump_task = asyncio.create_task(pty_to_ws())
    try:
        while True:
            msg = await ws.receive()
            if msg.get("type") == "websocket.disconnect":
                break
            if "bytes" in msg and msg["bytes"] is not None:
                try:
                    os.write(fd, msg["bytes"])
                except OSError:
                    break
                continue
            text = msg.get("text")
            if not text:
                continue
            try:
                obj = json.loads(text)
            except Exception:
                try:
                    os.write(fd, text.encode("utf-8"))
                except OSError:
                    break
                continue
            t = obj.get("type")
            if t == "input":
                try:
                    os.write(fd, (obj.get("data") or "").encode("utf-8"))
                except OSError:
                    break
            elif t == "resize":
                _set_size(int(obj.get("rows", 30)), int(obj.get("cols", 100)))
    except WebSocketDisconnect:
        pass
    except Exception:
        pass
    finally:
        closed = True
        try:
            os.kill(pid, signal.SIGHUP)
        except Exception:
            pass
        try:
            os.close(fd)
        except Exception:
            pass
        try:
            os.waitpid(pid, os.WNOHANG)
        except Exception:
            pass
        pump_task.cancel()
        try:
            await ws.close()
        except Exception:
            pass


@app.get("/")
def serve_index():
    return FileResponse(os.path.join(_ROOT, "index.html"))

@app.get("/index.html")
def serve_index_explicit():
    return FileResponse(os.path.join(_ROOT, "index.html"))

if __name__ == "__main__":
    import uvicorn
    # reload=True 需要传 "module:attr" 字符串形式,让 uvicorn 自己 import,这样改 .py 文件可热重载
    uvicorn.run("server:app", host="127.0.0.1", port=8000, reload=True)
